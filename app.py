from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from flask_pymongo import PyMongo
from bson.objectid import ObjectId
from datetime import datetime, timedelta
from werkzeug.security import generate_password_hash, check_password_hash
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
from email.message import EmailMessage
import smtplib
import ssl
import os
import pandas as pd
import io
from dotenv import load_dotenv 
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_talisman import Talisman
from flask_cors import CORS
load_dotenv()

app = Flask(__name__)
CORS(app)
Talisman(app, content_security_policy=None)  # CSP disabled for now to avoid breaking SPA scripts, can be hardened later
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://",
)

# --- CONFIGURATION ---
mongo_uri = os.environ.get("MONGO_URI")
if mongo_uri:
    mongo_uri = mongo_uri.strip()
    # Handle accidental label inclusion (e.g., "MONGO_URI: mongodb+srv://...")
    if mongo_uri.lower().startswith("mongo_uri:"):
        mongo_uri = mongo_uri[10:].strip()
    elif mongo_uri.lower().startswith("mongodb:"):
        # This is already a valid scheme, skip
        pass
    elif ":" in mongo_uri and not mongo_uri.startswith("mongodb"):
        # If there's a colon but it doesn't look like a scheme, it might be a label we missed
        parts = mongo_uri.split(":", 1)
        if len(parts) > 1 and "mongodb" in parts[1].lower():
            mongo_uri = parts[1].strip()
else:
    # Fallback for local dev if .env is missing, but Render will need this set
    print("WARNING: MONGO_URI environment variable is not set. Database connection will fail.")
    mongo_uri = "mongodb://localhost:27017/hospital_management" 
app.config["MONGO_URI"] = mongo_uri

secret_key = os.environ.get("SECRET_KEY", "06e4b4738ab81f94277a7216b5e79fb24b339f28a6a131391d8d6f8f0a295dc1")
app.config["SECRET_KEY"] = secret_key
app.config["GMAIL_USER"] = os.environ.get("GMAIL_USER")
app.config["GMAIL_APP_PASSWORD"] = os.environ.get("GMAIL_APP_PASSWORD")
app.config["PASSWORD_RESET_EXPIRY_MINUTES"] = int(os.environ.get("PASSWORD_RESET_EXPIRY_MINUTES", "30"))

try:
    mongo = PyMongo(app)
    # Trigger a simple operation to verify connection
    with app.app_context():
        # Using a timeout to ensure startup doesn't hang indefinitely
        mongo.cx.admin.command('ping')
    print("SUCCESS: Connected to MongoDB.")
except Exception as e:
    print(f"CRITICAL: MongoDB initialization failed: {type(e).__name__} - {e}")
    mongo = None

serializer = URLSafeTimedSerializer(app.config["SECRET_KEY"])

# --- HELPER: DATABASE CHECK & INITIAL SETUP ---
def check_db():
    if mongo is None or mongo.db is None:
        print("Database connection failed or not initialized.")
        return False
    return True

def clean_input_data(data):
    """Strip trailing and leading spaces from string values in a dictionary."""
    if not isinstance(data, dict):
        return data
    
    cleaned = {}
    for key, value in data.items():
        if isinstance(value, str):
            cleaned[key] = value.strip()
        elif isinstance(value, dict):
            cleaned[key] = clean_input_data(value)
        elif isinstance(value, list):
            cleaned[key] = [clean_input_data(item) if isinstance(item, dict) else item.strip() if isinstance(item, str) else item for item in value]
        else:
            cleaned[key] = value
    return cleaned

def ensure_initial_admin():
    """Checks for and creates the default admin user 'ImranSaab' on first run."""
    if check_db():
        if mongo.db.users.count_documents({}) == 0:
            # Create ImranSaab as the Admin
            admin_user = {
                'username': 'ImranSaab',
                'password': generate_password_hash('password123'),
                'role': 'Admin',
                'name': 'Imran Khan (Admin)',
                'email': os.environ.get('ADMIN_EMAIL', 'admin@example.com').strip().lower(),
                'created_at': datetime.now()
            }
            mongo.db.users.insert_one(admin_user)
            print("Initial Admin user 'ImranSaab' created.")

def create_indices():
    """Ensure essential database indices exist for performance."""
    if check_db():
        try:
            # Users index
            mongo.db.users.create_index([("username", 1)], unique=True)
            
            # Patients indices
            mongo.db.patients.create_index([("admissionDate", -1)])
            mongo.db.patients.create_index([("isDischarged", 1)])
            
            # Canteen Sales indices
            mongo.db.canteen_sales.create_index([("patient_id", 1)])
            mongo.db.canteen_sales.create_index([("date", -1)])
            
            # Expenses indices
            mongo.db.expenses.create_index([("date", -1)])
            mongo.db.expenses.create_index([("category", 1)])
            
            print("Database indices verified/created.")
        except Exception as e:
            print(f"Error creating indices: {e}")

# Run initial setup outside of request context
with app.app_context():
    ensure_initial_admin()
    create_indices()


def normalize_email(value):
    return value.strip().lower() if isinstance(value, str) else value


def send_password_reset_email(to_email, username, token):
    """Send a password reset email using Gmail SMTP credentials."""
    gmail_user = app.config.get("GMAIL_USER")
    gmail_pass = app.config.get("GMAIL_APP_PASSWORD")

    if not gmail_user or not gmail_pass:
        print("Gmail credentials missing; cannot send password reset email.")
        return False

    base_url = url_for('index', _external=True)
    connector = '&' if '?' in base_url else '?'
    reset_link = f"{base_url}{connector}reset_token={token}"
    expires_in = app.config.get("PASSWORD_RESET_EXPIRY_MINUTES", 30)

    message = EmailMessage()
    message["Subject"] = "Reset your PRO account password"
    message["From"] = gmail_user
    message["To"] = to_email
    message.set_content(
        f"Hello {username},\n\n"
        "We received a request to reset your password. "
        f"Use the link below to set a new password (valid for {expires_in} minutes).\n\n"
        f"{reset_link}\n\n"
        "If you did not request this, you can safely ignore this email."
    )

    try:
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
            server.login(gmail_user, gmail_pass)
            server.send_message(message)
        return True
    except Exception as e:
        print(f"Failed to send reset email: {e}")
        return False


# --- AUTHENTICATION ROUTES ---

def login_required(f):
    def wrapper(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({"error": "Unauthorized"}), 401
        return f(*args, **kwargs)
    wrapper.__name__ = f.__name__
    return wrapper

def role_required(roles):
    def decorator(f):
        @login_required
        def wrapper(*args, **kwargs):
            if not check_db(): return jsonify({"error": "Database not initialized"}), 500
            user = mongo.db.users.find_one({"_id": ObjectId(session['user_id'])})
            if user and user.get('role') in roles:
                return f(*args, **kwargs)
            return jsonify({"error": "Access Denied"}), 403
        wrapper.__name__ = f.__name__
        return wrapper
    return decorator

def calculate_prorated_fee(monthly_fee, days_elapsed):
    """
    Calculate prorated fee based purely on days elapsed.
    Formula: (monthly_fee / 30) * days_elapsed.
    This ensures patients are charged fairly per day based on their set monthly fee,
    calculating properly for both short stays and long stays.
    """
    try:
        # Parse monthly_fee to handle string values with commas
        if isinstance(monthly_fee, str):
            monthly_fee = int(monthly_fee.replace(',', '') or '0')
        else:
            monthly_fee = int(monthly_fee or 0)
        
        per_day_rate = monthly_fee / 30.0
        return int(per_day_rate * max(days_elapsed, 1))  # At least 1 day charge
    except (ValueError, TypeError):
        return 0


# ============================================================
# FINANCIAL SYSTEM LOGIC OVERVIEW:
# ============================================================
# 
# The system tracks patient finances through multiple components:
#
# 1. PATIENT CHARGES (Calculated):
#    - Monthly Fee: Stored per patient, prorated after 90 days
#    - Canteen Sales: Aggregated from canteen_sales collection
#    - Laundry: One-time charge added at discharge (if laundryStatus=True)
#    
# 2. PAYMENTS (Tracked):
#    - receivedAmount: Cumulative payments stored in patient record
#    - Payment History: Individual payments logged in expenses collection
#      (type='incoming', category='Patient Fee', auto=True)
#
# 3. BALANCE CALCULATION:
#    Balance Due = (Fee + Canteen + Laundry) - Received Amount
#
# 4. DASHBOARD METRICS:
#    - Total Expected Balance: Sum of all positive balances from active patients
#    - This shows total money owed to the facility
#
# 5. EXPENSES TRACKING:
#    - Manual Income: Recorded in expenses (type='incoming')  
#    - Manual Outgoing: Recorded in expenses (type='outgoing')
#    - Patient payments are auto-recorded but NOT double-counted in summaries
#
# 6. OVERHEADS TRACKING:
#    - Monthly daily expense tracking (kitchen, canteen, others, advances, income)
#    - Canteen column auto-syncs with canteen_sales collection
#    - Shows daily profit/loss calculations
#
# 7. DATA CONSISTENCY:
#    - Canteen totals: Aggregated from canteen_sales using patient_id
#    - Payments: receivedAmount must match sum of payment history
#    - All financial fields stored as strings with commas, parsed as integers
# ============================================================

@app.route('/')
def index():
    # Frontend handles redirection to login if session is missing.
    return render_template('index.html')

@app.route('/api/auth/login', methods=['POST'])
@limiter.limit("5 per minute")
def login():
    if not check_db(): return jsonify({"error": "Database error"}), 500
    data = clean_input_data(request.json)
    user = mongo.db.users.find_one({"username": data['username']})
    
    if user and check_password_hash(user['password'], data['password']):
        session['user_id'] = str(user['_id'])
        session['username'] = user['username']
        session['role'] = user['role']
        return jsonify({
            "message": "Login successful",
            "username": user['username'],
            "role": user['role'],
            "name": user.get('name', user['username']),
            "user_id": str(user['_id'])
        })
    return jsonify({"error": "Invalid credentials"}), 401


@app.route('/api/auth/forgot', methods=['POST'])
def forgot_password():
    """Initiate password reset by emailing a time-bound token."""
    if not check_db():
        return jsonify({"error": "Database error"}), 500

    data = clean_input_data(request.json or {})
    username = data.get('username')
    email = normalize_email(data.get('email'))

    if not username or not email:
        return jsonify({"error": "Username and email are required"}), 400

    if not app.config.get("GMAIL_USER") or not app.config.get("GMAIL_APP_PASSWORD"):
        return jsonify({"error": "Email service not configured"}), 500

    user = mongo.db.users.find_one({"username": username})
    if not user:
        return jsonify({"error": "No account found for that username."}), 404

    registered_email = normalize_email(user.get('email'))
    if not registered_email:
        return jsonify({"error": "No email is set for this account. Contact an admin."}), 400

    if registered_email != email:
        return jsonify({"error": "Username and email do not match our records."}), 400

    token = serializer.dumps({"user_id": str(user['_id']), "email": registered_email}, salt="password-reset")
    sent = send_password_reset_email(registered_email, user.get('name', user['username']), token)
    if not sent:
        return jsonify({"error": "Could not send reset email. Please try again or contact support."}), 500

    return jsonify({"message": "Reset email sent to your registered address."})


@app.route('/api/auth/reset', methods=['POST'])
def reset_password():
    """Reset password using a token delivered via email."""
    if not check_db():
        return jsonify({"error": "Database error"}), 500

    data = clean_input_data(request.json or {})
    token = data.get('token')
    new_password = data.get('new_password')

    if not token or not new_password:
        return jsonify({"error": "Token and new password are required"}), 400

    try:
        payload = serializer.loads(
            token,
            salt="password-reset",
            max_age=app.config.get("PASSWORD_RESET_EXPIRY_MINUTES", 30) * 60
        )
    except SignatureExpired:
        return jsonify({"error": "Reset link expired"}), 400
    except BadSignature:
        return jsonify({"error": "Invalid reset token"}), 400

    user_id = payload.get('user_id')
    email = normalize_email(payload.get('email'))
    if not user_id:
        return jsonify({"error": "Invalid reset token"}), 400

    user = mongo.db.users.find_one({"_id": ObjectId(user_id)})
    if not user or (email and normalize_email(user.get('email')) != email):
        return jsonify({"error": "Invalid reset token"}), 400

    new_password_hash = generate_password_hash(new_password)
    mongo.db.users.update_one({'_id': ObjectId(user_id)}, {'$set': {'password': new_password_hash}})

    return jsonify({"message": "Password has been reset successfully"})

@app.route('/api/auth/logout', methods=['POST'])
def logout():
    session.pop('user_id', None)
    session.pop('username', None)
    session.pop('role', None)
    return jsonify({"message": "Logged out"})

@app.route('/api/auth/session', methods=['GET'])
def check_session():
    if 'user_id' in session:
        return jsonify({
            "is_logged_in": True,
            "username": session.get('username'),
            "role": session.get('role'),
            "user_id": session.get('user_id')
        })
    return jsonify({"is_logged_in": False})

# --- USER MANAGEMENT (ADMIN ONLY) ---
@app.route('/api/users', methods=['GET'])
@role_required(['Admin'])
def get_users():
    if not check_db(): return jsonify([])
    users_cursor = mongo.db.users.find({}, {'password': 0})
    users = [{**u, '_id': str(u['_id'])} for u in users_cursor]
    return jsonify(users)

@app.route('/api/users', methods=['POST'])
@role_required(['Admin'])
def create_user():
    if not check_db(): return jsonify({"error": "Database error"}), 500
    data = clean_input_data(request.json)
    if not all(k in data for k in ['username', 'password', 'role', 'name', 'email']):
        return jsonify({"error": "Missing fields"}), 400

    data['email'] = normalize_email(data.get('email'))
    if not data['email']:
        return jsonify({"error": "Valid email required"}), 400
    
    if mongo.db.users.find_one({"username": data['username']}):
        return jsonify({"error": "Username already exists"}), 409

    if mongo.db.users.find_one({"email": data['email']}):
        return jsonify({"error": "Email already exists"}), 409

    data['password'] = generate_password_hash(data['password'])
    data['created_at'] = datetime.now()
    try:
        result = mongo.db.users.insert_one(data)
        return jsonify({"message": "User created", "id": str(result.inserted_id)}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500
@app.route('/api/users/<id>', methods=['DELETE'])
@role_required(['Admin'])
def delete_user(id):
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        # Prevent deleting the main admin
        user = mongo.db.users.find_one({'_id': ObjectId(id)})
        if user and user.get('username') == 'ImranSaab':
            return jsonify({"error": "Main admin cannot be deleted"}), 403
            
        mongo.db.users.delete_one({'_id': ObjectId(id)})
        return jsonify({"message": "User deleted"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# @app.route('/api/patients/<id>', methods=['DELETE'])
# @role_required(['Admin'])
# def delete_patient(id):
#     if not check_db(): return jsonify({"error": "Database error"}), 500
#     try:
#         mongo.db.patients.delete_one({'_id': ObjectId(id)})
#         return jsonify({"message": "Patient deleted"})
#     except Exception as e:
#         return jsonify({"error": str(e)}), 500


@app.route('/api/users/change_password', methods=['POST'])
@login_required
def change_password():
    if not check_db(): return jsonify({"error": "Database error"}), 500
    data = clean_input_data(request.json)
    user_id = session['user_id']
    
    try:
        # User is changing their own password
        user = mongo.db.users.find_one({"_id": ObjectId(user_id)})
        if not user or not check_password_hash(user['password'], data['old_password']):
            return jsonify({"error": "Invalid old password"}), 401
        
        new_password_hash = generate_password_hash(data['new_password'])
        mongo.db.users.update_one({'_id': ObjectId(user_id)}, {'$set': {'password': new_password_hash}})
        return jsonify({"message": "Password updated successfully"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# --- DASHBOARD METRICS ---
@app.route('/api/dashboard', methods=['GET'])
@login_required
def get_dashboard_metrics():
    if not check_db(): return jsonify({"error": "Database error"}), 500
    
    today = datetime.now()
    start_of_month = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    if today.month == 12:
        end_of_month = today.replace(year=today.year + 1, month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        end_of_month = today.replace(month=today.month + 1, day=1, hour=0, minute=0, second=0, microsecond=0)
    
    try:
        # 1. Basic Counts
        total_patients = mongo.db.patients.count_documents({'isDischarged': {'$ne': True}})
        admissions_this_month = mongo.db.patients.count_documents({
            'admissionDate': {'$gte': start_of_month.isoformat(), '$lt': end_of_month.isoformat()}
        })
        discharges_this_month = mongo.db.patients.count_documents({
            'isDischarged': True,
            'dischargeDate': {'$gte': start_of_month.isoformat(), '$lt': end_of_month.isoformat()}
        })
        
        # 2. Total Expected Incoming (Remaining Balance Calculation)
        # Fetch ALL canteen sales first to create a map
        # Note: We group by patient_id. We convert ObjectId to string for easy matching.
        all_canteen_sales = list(mongo.db.canteen_sales.find())
        canteen_map = {}
        
        for sale in all_canteen_sales:
            # Handle patient_id whether it's ObjectId or String
            pid = str(sale.get('patient_id', ''))
            amount = int(sale.get('amount', 0))
            if pid:
                canteen_map[pid] = canteen_map.get(pid, 0) + amount

        # Fetch Active Patients
        active_patients = list(mongo.db.patients.find({'isDischarged': {'$ne': True}}))
        
        total_expected_balance = 0
        
        # Calculate total expected balance from active patients (fee + canteen + laundry - received)
        for patient in active_patients:
            try:
                pid = str(patient['_id'])
                
                # Calculate days elapsed for prorated fee
                admission_date = patient.get('admissionDate')
                days_elapsed = 0
                if admission_date:
                    try:
                        if isinstance(admission_date, str):
                            admission_dt = datetime.fromisoformat(admission_date.replace('Z', '+00:00'))
                        else:
                            admission_dt = admission_date
                        days_diff = (datetime.now() - admission_dt).days
                        days_elapsed = max(0, days_diff)
                    except:
                        pass
                
                # Get prorated fee
                fee_str = patient.get('monthlyFee', '0') or '0'
                fee = calculate_prorated_fee(fee_str, days_elapsed)
                
                # Get canteen total
                canteen = canteen_map.get(pid, 0)
                
                # Get laundry (one-time charge for discharge)
                laundry = patient.get('laundryAmount', 0) if patient.get('laundryStatus', False) else 0
                
                # Get received amount
                received_str = str(patient.get('receivedAmount', '0')).replace(',', '')
                received = int(received_str or '0')
                
                # Calculate remaining balance
                balance = fee + canteen + laundry - received
                total_expected_balance += max(0, balance)  # Only count positive balances
            except (ValueError, TypeError) as e:
                print(f"Dashboard calculation error for patient {patient.get('name')}: {e}")
                pass

        # 3. Canteen Sales This Month (KPI Card)
        pipeline_month = [
            {'$match': {'date': {'$gte': start_of_month, '$lt': end_of_month}}},
            {'$group': {'_id': None, 'total_sales': {'$sum': '$amount'}}}
        ]
        canteen_month_res = list(mongo.db.canteen_sales.aggregate(pipeline_month))
        total_canteen_sales_this_month = canteen_month_res[0]['total_sales'] if canteen_month_res else 0
        
        # 4. Total Expenses This Month (KPI Card)
        pipeline_expenses = [
            {'$match': {
                'type': 'outgoing',
                'date': {'$gte': start_of_month, '$lt': end_of_month}
            }},
            {'$group': {'_id': None, 'total': {'$sum': '$amount'}}}
        ]
        expenses_res = list(mongo.db.expenses.aggregate(pipeline_expenses))
        total_expenses_this_month = expenses_res[0]['total'] if expenses_res else 0
        
        # 5. Psychology Sessions Today
        today_start = today.replace(hour=0, minute=0, second=0, microsecond=0)
        today_end = today_start + timedelta(days=1)
        total_psych_sessions_today = mongo.db.psych_sessions.count_documents({
            'date': {'$gte': today_start, '$lt': today_end}
        })
        
        return jsonify({
            'totalPatients': total_patients,
            'admissionsThisMonth': admissions_this_month,
            'dischargesThisMonth': discharges_this_month,
            'totalExpectedBalance': total_expected_balance,
            'totalCanteenSalesThisMonth': total_canteen_sales_this_month,
            'totalExpensesThisMonth': total_expenses_this_month,
            'totalPsychSessionsToday': total_psych_sessions_today
        })
    except Exception as e:
        print(f"DB Metric Error: {e}")
        return jsonify({"error": str(e)}), 500


# DEBUG endpoint to inspect database
@app.route('/api/debug/dashboard', methods=['GET'])
@login_required
def debug_dashboard():
    """Debug endpoint to show raw data used in dashboard calculations"""
    if not check_db(): return jsonify({"error": "Database error"}), 500
    
    today = datetime.now()
    start_of_month = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    try:
        # Get all patients with fees
        patients = list(mongo.db.patients.find({'isDischarged': {'$ne': True}}))
        patient_data = []
        for p in patients:
            try:
                fee = int(p.get('monthlyFee', '0').replace(',', ''))
                patient_data.append({
                    'name': p.get('name'),
                    'monthlyFee_raw': p.get('monthlyFee'),
                    'monthlyFee_parsed': fee
                })
            except ValueError:
                patient_data.append({
                    'name': p.get('name'),
                    'monthlyFee_raw': p.get('monthlyFee'),
                    'monthlyFee_parsed': 'ERROR'
                })
        
        # Get canteen sales this month
        canteen_pipeline = [
            {'$match': {'date': {'$gte': start_of_month}}},
            {'$group': {'_id': None, 'total': {'$sum': '$amount'}, 'count': {'$sum': 1}}}
        ]
        canteen_data = list(mongo.db.canteen_sales.aggregate(canteen_pipeline))
        
        # Get all canteen sales for context
        all_canteen = list(mongo.db.canteen_sales.find().sort('date', -1).limit(5))
        canteen_sample = [{
            'date': str(c.get('date')),
            'amount': c.get('amount'),
            'item': c.get('item')
        } for c in all_canteen]
        
        return jsonify({
            'currentMonth': f"{today.year}-{today.month:02d}",
            'startOfMonth': str(start_of_month),
            'totalPatients': len(patients),
            'patientsWithFees': patient_data,
            'canteenThisMonth': canteen_data,
            'canteenSample': canteen_sample
        })
    except Exception as e:
        print(f"Debug error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/dashboard/admissions', methods=['GET'])
@login_required
def get_month_admissions():
    """Return detailed admissions for the current month."""
    if not check_db():
        return jsonify({"error": "Database error"}), 500

    today = datetime.now()
    start_of_month = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    try:
        cursor = mongo.db.patients.find({'created_at': {'$gte': start_of_month}})
        admissions = []
        for p in cursor:
            admissions.append({
                'id': str(p.get('_id')),
                'name': p.get('name', ''),
                'admissionDate': p.get('admissionDate', ''),
                'created_at': p.get('created_at').isoformat() if p.get('created_at') else ''
            })
        return jsonify(admissions)
    except Exception as e:
        print(f"Admissions list error: {e}")
        return jsonify({"error": str(e)}), 500

# --- PATIENT API UPDATES ---

@app.route('/api/patients', methods=['GET'])
@login_required
def get_patients():
    if not check_db(): return jsonify([])
    try:
        patients_cursor = mongo.db.patients.find()
        
        # Aggregate total canteen spending for all patients
        canteen_totals_agg = list(mongo.db.canteen_sales.aggregate([
            {'$match': {
                '$or': [
                    {'entry_type': {'$exists': False}},
                    {'entry_type': {'$ne': 'other'}}
                ]
            }},
            {'$group': {'_id': '$patient_id', 'total': {'$sum': '$amount'}}}
        ]))
        canteen_totals_map = {str(item['_id']): item['total'] for item in canteen_totals_agg}
        
        patients = []
        for p in patients_cursor:
            patient_id = str(p['_id'])
            p['_id'] = patient_id
            # Ensure monthlyFee is present for canteen view logic
            p['monthlyFee'] = p.get('monthlyFee', '0')
            p['photo1'] = p.get('photo1', '')
            p['photo2'] = p.get('photo2', '')
            p['photo3'] = p.get('photo3', '')
            p['isDischarged'] = p.get('isDischarged', False)
            p['dischargeDate'] = p.get('dischargeDate')
            
            # Include canteen spending as separate field
            p['canteenSpent'] = canteen_totals_map.get(patient_id, 0)
            
            patients.append(p)
        return jsonify(patients)
    except Exception as e:
        print(f"DB Fetch Error: {e}")
        return jsonify([])

@app.route('/api/patients', methods=['POST'])
@role_required(['Admin', 'Doctor']) # Only Admin/Doctor can admit
def add_patient():
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        data = clean_input_data(request.json)
        data['created_at'] = datetime.now()
        data['notes'] = [] # General Notes (Legacy)
        data['monthlyFee'] = data.get('monthlyFee', '0')
        data['monthlyAllowance'] = data.get('monthlyAllowance', '3000') # Default allowance
        data['receivedAmount'] = data.get('receivedAmount', '0')  # New field
        data['drug'] = data.get('drug', '')  # New field
        data['photo1'] = data.get('photo1', '')
        data['photo2'] = data.get('photo2', '')
        data['photo3'] = data.get('photo3', '')
        data['isDischarged'] = data.get('isDischarged', False)
        data['dischargeDate'] = data.get('dischargeDate')
        
        # Laundry fields (one-time charge added to final discharge bill)
        data['laundryStatus'] = data.get('laundryStatus', False)  # Boolean: whether laundry service is enabled
        if data['laundryStatus']:
            data['laundryAmount'] = int(data.get('laundryAmount', 3500))  # Default 3500 if enabled (one-time charge)
        else:
            data['laundryAmount'] = 0  # 0 if not enabled
        
        result = mongo.db.patients.insert_one(data)
        patient_id = str(result.inserted_id)

        # Auto-log initial payment as an expense
        try:
            initial_received = int(str(data.get('receivedAmount', '0')).replace(',', ''))
            if initial_received > 0:
                mongo.db.expenses.insert_one({
                    'type': 'incoming',
                    'amount': initial_received,
                    'category': 'Patient Fee',
                    'note': f"Initial Advance from {data.get('name')} (Admission)",
                    'payment_method': 'Cash/Initial',
                    'patient_id': patient_id,
                    'date': datetime.now(),
                    'recorded_by': session.get('username', 'Admin'),
                    'auto': True
                })
        except: pass

        return jsonify({"message": "Success", "id": patient_id}), 201
    except Exception as e:
        print(f"DB Insert Error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/patients/<id>', methods=['PUT'])
@role_required(['Admin', 'Doctor'])
def update_patient(id):
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        data = clean_input_data(request.json)
        if '_id' in data: del data['_id']
        
        # Allow all users to update all fields. (Restrictions removed as per user request)
        
        mongo.db.patients.update_one({'_id': ObjectId(id)}, {'$set': data})
        return jsonify({"message": "Updated"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/patients/<id>', methods=['DELETE'])
@role_required(['Admin'])
def delete_patient(id):
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        # Delete the patient
        result = mongo.db.patients.delete_one({'_id': ObjectId(id)})
        if result.deleted_count > 0:
            # Also delete associated records (session notes and medical records)
            mongo.db.patient_records.delete_many({'patient_id': id})
            return jsonify({"message": "Patient deleted successfully"}), 200
        else:
            return jsonify({"error": "Patient not found"}), 404
    except Exception as e:
        print(f"Delete Error: {e}")
        return jsonify({"error": str(e)}), 500

# --- NEW PATIENT RECORD APIS (SESSION NOTES & MEDICAL RECORDS) ---

@app.route('/api/patients/<patient_id>/session_note', methods=['POST'])
@role_required(['Admin', 'Psychologist'])
def add_session_note(patient_id):
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        data = clean_input_data(request.json)
        note = {
            'text': data['text'],
            'type': 'session_note',
            'date': datetime.now(),
            'recorded_by': session.get('username', 'System'),
            'patient_id': ObjectId(patient_id)
        }
        result = mongo.db.patient_records.insert_one(note)
        return jsonify({"message": "Session note added", "id": str(result.inserted_id)}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/patients/<patient_id>/medical_record', methods=['POST'])
@role_required(['Admin', 'Doctor'])
def add_medical_record(patient_id):
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        data = clean_input_data(request.json)
        record = {
            'title': data['title'],
            'details': data['details'],
            'type': 'medical_record',
            'date': datetime.now(),
            'recorded_by': session.get('username', 'System'),
            'patient_id': ObjectId(patient_id)
        }
        result = mongo.db.patient_records.insert_one(record)
        return jsonify({"message": "Medical record added", "id": str(result.inserted_id)}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500
        
@app.route('/api/patients/<patient_id>/records', methods=['GET'])
@login_required
def get_patient_records(patient_id):
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        records_cursor = mongo.db.patient_records.find({'patient_id': ObjectId(patient_id)}).sort('date', -1)
        records = []
        for r in records_cursor:
            r['_id'] = str(r['_id'])
            r['patient_id'] = str(r['patient_id'])
            r['date'] = r['date'].isoformat()
            records.append(r)
        return jsonify(records)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# --- CANTEEN APIS ---

@app.route('/api/canteen/sales', methods=['POST'])
@role_required(['Admin', 'Canteen'])
def record_canteen_sale():
    if not check_db(): return jsonify({"error": "Database error"}), 500
    data = clean_input_data(request.json)
    if not all(k in data for k in ['patient_id', 'amount', 'item']):
        return jsonify({"error": "Missing fields"}), 400
    
    try:
        # Convert amount to integer
        data['amount'] = int(data['amount'])
        
        # Get the date, default to today if not provided
        sale_date = data.get('date')
        if sale_date:
            sale_date = datetime.fromisoformat(sale_date.replace('Z', '+00:00'))
        else:
            sale_date = datetime.now()
        
        sale = {
            'patient_id': ObjectId(data['patient_id']),
            'item': data['item'],
            'amount': data['amount'],
            'date': sale_date,
            'recorded_by': session.get('username', 'Canteen Staff')
        }
        result = mongo.db.canteen_sales.insert_one(sale)
        return jsonify({"message": "Sale recorded", "id": str(result.inserted_id)}), 201
    except ValueError:
        return jsonify({"error": "Amount must be a number"}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/canteen/sales/breakdown', methods=['GET'])
@role_required(['Admin', 'Canteen'])
def get_canteen_breakdown():
    if not check_db(): return jsonify({"error": "Database error"}), 500
    
    today = datetime.now()
    start_of_month = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    # Calculate days in current month
    if today.month == 12:
        next_month = today.replace(year=today.year + 1, month=1, day=1)
    else:
        next_month = today.replace(month=today.month + 1, day=1)
    days_in_month = (next_month - start_of_month).days
    
    try:
        def _is_discharged_flag(raw_val):
            if isinstance(raw_val, bool):
                return raw_val
            return str(raw_val).strip().lower() in ('true', '1', 'yes')

        # 1. Fetch all patients with ID, Name, Allowance AND isDischarged
        patients_cursor = mongo.db.patients.find({}, {
            'name': 1, 'monthlyAllowance': 1, 'isDischarged': 1
        })
        
        patients_map = {
            str(p['_id']): {
                'name': p['name'], 
                'allowance': p.get('monthlyAllowance', '0'), 
                'sales': 0,
                'isDischarged': p.get('isDischarged', False)
            } 
            for p in patients_cursor
            if not _is_discharged_flag(p.get('isDischarged', False))
        }
        
        # 2. Calculate monthly sales per patient
        pipeline = [
            {'$match': {'date': {'$gte': start_of_month}}},
            {'$group': {'_id': '$patient_id', 'total_sales': {'$sum': '$amount'}}}
        ]
        sales_breakdown = list(mongo.db.canteen_sales.aggregate(pipeline))
        
        # 3. Merge data
        for sale in sales_breakdown:
            p_id = str(sale['_id'])
            if p_id in patients_map:
                patients_map[p_id]['sales'] = sale['total_sales']
        
        # Format output
        breakdown_list = []
        for p_id, data in patients_map.items():
            try:
                sales = data['sales']
                monthly_allowance = int(data['allowance'].replace(',', ''))
                # Calculate daily allowance
                daily_allowance = monthly_allowance / days_in_month if days_in_month > 0 else 0
                balance = monthly_allowance - sales
            except ValueError:
                sales = data['sales']
                monthly_allowance = 0
                daily_allowance = 0
                balance = -sales
                
            breakdown_list.append({
                'id': p_id,
                'name': data['name'],
                'monthlyAllowance': data['allowance'],
                'dailyAllowance': round(daily_allowance, 2),
                'monthlySales': sales,
                'remainingBalance': balance,
                'isDischarged': data['isDischarged']
            })
            
        return jsonify(breakdown_list)
    except Exception as e:
        print(f"Canteen Breakdown Error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/canteen/daily-sheet', methods=['GET'])
@role_required(['Admin', 'Canteen'])
def get_daily_canteen_sheet():
    """Get daily canteen sheet for today with all active patients"""
    if not check_db(): return jsonify({"error": "Database error"}), 500
    
    try:
        # Get query parameter for date, default to today
        date_str = request.args.get('date')
        if date_str:
            target_date = datetime.fromisoformat(date_str)
        else:
            target_date = datetime.now()
        
        # Set time range for the target day
        start_of_day = target_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_of_day = target_date.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        # Fetch all active patients
        patients_cursor = mongo.db.patients.find(
            {'isDischarged': {'$ne': True}},
            {'name': 1, 'monthlyAllowance': 1}
        ).sort('name', 1)
        
        # Get today's sales for each patient
        pipeline = [
            {'$match': {'date': {'$gte': start_of_day, '$lte': end_of_day}}},
            {'$group': {
                '_id': '$patient_id',
                'items': {'$push': {'item': '$item', 'amount': '$amount'}},
                'total': {'$sum': '$amount'}
            }}
        ]
        daily_sales = {str(s['_id']): s for s in mongo.db.canteen_sales.aggregate(pipeline)}
        
        # Build sheet
        sheet = []
        for p in patients_cursor:
            p_id = str(p['_id'])
            sales_data = daily_sales.get(p_id, {'items': [], 'total': 0})
            
            sheet.append({
                'id': p_id,
                'name': p['name'],
                'dailyAllowance': p.get('monthlyAllowance', '0'),
                'todayItems': sales_data['items'],
                'todayTotal': sales_data['total']
            })
        
        return jsonify({
            'date': target_date.strftime('%Y-%m-%d'),
            'patients': sheet
        })
    except Exception as e:
        print(f"Daily Sheet Error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/canteen/sales/history', methods=['GET'])
@role_required(['Admin'])
def get_canteen_sales_history():
    """Get detailed canteen sales history - Admin only"""
    if not check_db(): return jsonify({"error": "Database error"}), 500
    
    try:
        patient_id = request.args.get('patient_id')
        
        query = {}
        if patient_id:
            query['patient_id'] = ObjectId(patient_id)
        
        # Get sales with patient names
        sales_cursor = mongo.db.canteen_sales.find(query).sort('date', -1).limit(100)
        
        sales_list = []
        for sale in sales_cursor:
            # Get patient name
            patient = mongo.db.patients.find_one({'_id': sale['patient_id']}, {'name': 1})
            
            sales_list.append({
                'id': str(sale['_id']),
                'patient_id': str(sale['patient_id']),
                'patient_name': patient['name'] if patient else 'Unknown',
                'item': sale.get('item', ''),
                'amount': sale.get('amount', 0),
                'date': sale['date'].isoformat() if sale.get('date') else '',
                'recorded_by': sale.get('recorded_by', '')
            })
        
        return jsonify(sales_list)
    except Exception as e:
        print(f"Sales History Error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/canteen/monthly-table', methods=['GET'])
@role_required(['Admin', 'Canteen'])
def get_canteen_monthly_table():
    """Get monthly canteen table data with daily columns"""
    if not check_db(): return jsonify({"error": "Database error"}), 500
    
    try:
        def _is_discharged_flag(raw_val):
            if isinstance(raw_val, bool):
                return raw_val
            return str(raw_val).strip().lower() in ('true', '1', 'yes')

        # Get month and year from query params, default to current month
        month = int(request.args.get('month', datetime.now().month))
        year = int(request.args.get('year', datetime.now().year))
        
        # Calculate start and end of requested month
        start_of_month = datetime(year, month, 1, 0, 0, 0)
        if month == 12:
            end_of_month = datetime(year + 1, 1, 1, 0, 0, 0)
        else:
            end_of_month = datetime(year, month + 1, 1, 0, 0, 0)
        
        days_in_month = (end_of_month - start_of_month).days
        
        # Get active patients only (exclude discharged from canteen tracker)
        raw_patients = list(mongo.db.patients.find({}, {
            'name': 1,
            'monthlyAllowance': 1,
            'isDischarged': 1,
            'admissionDate': 1
        }))
        patients_list = [p for p in raw_patients if not _is_discharged_flag(p.get('isDischarged', False))]
        patients_list.sort(key=lambda p: str(p.get('name', '')).lower())
        
        # Get manual old balance overrides for this month/year
        balance_overrides = {}
        overrides_cursor = mongo.db.canteen_balance_overrides.find({
            'month': month,
            'year': year
        })
        for override in overrides_cursor:
            balance_overrides[str(override['patient_id'])] = override['old_balance']
        
        if not patients_list:
            return jsonify({'month': month, 'year': year, 'daysInMonth': days_in_month, 'patients': []})
        
        patient_ids = [p['_id'] for p in patients_list]
        
        def _safe_int(raw_val: object) -> int:
            """Best-effort int conversion that strips non-digits."""
            try:
                cleaned = ''.join(ch for ch in str(raw_val or '0') if ch.isdigit() or ch == '-')
                return int(cleaned) if cleaned not in ('', '-') else 0
            except Exception:
                return 0
        
        # BATCH QUERY: Get all previous sales for all patients at once
        previous_sales_agg = list(mongo.db.canteen_sales.aggregate([
            {'$match': {
                'patient_id': {'$in': patient_ids},
                'date': {'$lt': start_of_month},
                '$or': [
                    {'entry_type': {'$exists': False}},
                    {'entry_type': {'$ne': 'other'}}
                ]
            }},
            {'$group': {'_id': '$patient_id', 'total': {'$sum': '$amount'}}}
        ]))
        previous_sales_map = {str(item['_id']): item['total'] for item in previous_sales_agg}
        
        # BATCH QUERY: Get all previous adjustments
        previous_adj_agg = list(mongo.db.canteen_sales.aggregate([
            {'$match': {
                'patient_id': {'$in': patient_ids},
                'date': {'$lt': start_of_month},
                'entry_type': 'other'
            }},
            {'$group': {'_id': '$patient_id', 'total': {'$sum': '$amount'}}}
        ]))
        previous_adj_map = {str(item['_id']): item['total'] for item in previous_adj_agg}
        
        # BATCH QUERY: Get all current month daily sales
        current_month_sales = list(mongo.db.canteen_sales.find({
            'patient_id': {'$in': patient_ids},
            'date': {'$gte': start_of_month, '$lt': end_of_month},
            '$or': [
                {'entry_type': {'$exists': False}},
                {'entry_type': {'$ne': 'other'}}
            ]
        }))
        
        # BATCH QUERY: Get all "other" entries for current month
        other_entries = list(mongo.db.canteen_sales.find({
            'patient_id': {'$in': patient_ids},
            'date': {'$gte': start_of_month, '$lt': end_of_month},
            'entry_type': 'other'
        }))
        other_map = {str(item['patient_id']): item['amount'] for item in other_entries}
        
        # BATCH QUERY: Get all-time totals for all patients
        all_time_agg = list(mongo.db.canteen_sales.aggregate([
            {'$match': {
                'patient_id': {'$in': patient_ids},
                '$or': [
                    {'entry_type': {'$exists': False}},
                    {'entry_type': {'$ne': 'other'}}
                ]
            }},
            {'$group': {'_id': '$patient_id', 'total': {'$sum': '$amount'}}}
        ]))
        all_time_map = {str(item['_id']): item['total'] for item in all_time_agg}
        
        patients_data = []

        for patient in patients_list:
            patient_id = patient['_id']
            patient_id_str = str(patient_id)
            patient_name = patient.get('name', 'Unknown')
            monthly_allowance = _safe_int(patient.get('monthlyAllowance', 0))
            is_discharged = patient.get('isDischarged', False)
            
            # Get data from batch queries
            previous_sales_total = previous_sales_map.get(patient_id_str, 0)
            previous_adjustments = previous_adj_map.get(patient_id_str, 0)
            
            # Old Balance = Sum of total canteen money used in all previous months since admission
            # This is simply the total of all canteen sales before the current viewing month
            calculated_balance = previous_sales_total
            
            # Check if there's a manual override for this patient's old balance
            old_balance = balance_overrides.get(patient_id_str, calculated_balance)
            has_override = patient_id_str in balance_overrides
            
            # Build daily entries from batch query results
            daily_entries = {}
            for sale in current_month_sales:
                if str(sale['patient_id']) == patient_id_str:
                    day = sale['date'].day
                    amount = sale.get('amount', 0)
                    if day in daily_entries:
                        daily_entries[day] += amount
                    else:
                        daily_entries[day] = amount
            
            # Get "other" amount from batch query
            other_amount = other_map.get(patient_id_str, 0)
            
            # Calculate Month Total (sum of daily entries + other)
            month_total = sum(daily_entries.values()) + other_amount
            
            # Get all-time total from batch query
            total_spent = all_time_map.get(patient_id_str, 0)
            
            patients_data.append({
                'id': str(patient_id),
                'name': patient_name,
                'oldBalance': old_balance,
                'calculatedBalance': calculated_balance,
                'hasManualOverride': has_override,
                'dailyEntries': daily_entries,
                'other': other_amount,
                'monthTotal': month_total,
                'total': total_spent,
                'isDischarged': is_discharged,
                'exceedsBalance': month_total > old_balance
            })
        
        return jsonify({
            'month': month,
            'year': year,
            'daysInMonth': days_in_month,
            'patients': patients_data
        })
    except Exception as e:
        print(f"Monthly Table Error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/canteen/old-balance', methods=['POST'])
@role_required(['Admin'])
def save_canteen_old_balance():
    """
    Save manual override for canteen old balance.
    
    IMPORTANT: This only affects the "Old Balance" display column in the monthly 
    canteen tracking table. It does NOT affect actual patient billing or financial 
    calculations. 
    
    The Old Balance is a budgeting/tracking feature that shows the allowance 
    available at the start of the month. Actual billing uses the sum of 
    canteen_sales entries, not this old balance field.
    """
    if not check_db(): return jsonify({"error": "Database error"}), 500
    
    data = clean_input_data(request.json)
    try:
        patient_id = data.get('patient_id')
        month = int(data.get('month'))
        year = int(data.get('year'))
        old_balance = int(data.get('old_balance', 0))
        
        # Upsert or delete the override
        if old_balance == 0:
            mongo.db.canteen_balance_overrides.delete_one({
                'patient_id': ObjectId(patient_id),
                'month': month,
                'year': year
            })
            return jsonify({"message": "Old balance override removed"})
        else:
            mongo.db.canteen_balance_overrides.update_one(
                {
                    'patient_id': ObjectId(patient_id),
                    'month': month,
                    'year': year
                },
                {
                    '$set': {
                        'old_balance': old_balance,
                        'updated_at': datetime.now(),
                        'updated_by': session.get('username')
                    }
                },
                upsert=True
            )
            return jsonify({"message": "Old balance updated"})
    except Exception as e:
        print(f"Save Old Balance Error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/canteen/daily-entry', methods=['POST'])
@role_required(['Admin', 'Canteen'])
def save_canteen_daily_entry():
    """Save or update a daily canteen entry"""
    if not check_db(): return jsonify({"error": "Database error"}), 500
    
    data = clean_input_data(request.json)
    if not all(k in data for k in ['patient_id', 'date', 'amount', 'entry_type']):
        return jsonify({"error": "Missing required fields"}), 400
    
    try:
        patient_id = ObjectId(data['patient_id'])
        entry_date = datetime.fromisoformat(data['date'].replace('Z', '+00:00'))
        amount = int(data['amount'])
        entry_type = data['entry_type']  # 'daily' or 'other'
        
        start_of_day = entry_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_of_day = start_of_day + timedelta(days=1)
        
        # Check if entry already exists within the day
        query = {
            'patient_id': patient_id,
            'date': {'$gte': start_of_day, '$lt': end_of_day}
        }
        if entry_type == 'other':
            query['entry_type'] = 'other'
        else:
            # Match explicitly 'daily' OR older legacy items that lacked entry_type
            query['$or'] = [{'entry_type': 'daily'}, {'entry_type': {'$exists': False}}]
            
        existing_entry = mongo.db.canteen_sales.find_one(query)
        
        # Role-based permission check
        user_role = session.get('role')
        username = session.get('username', 'Unknown')
        
        if existing_entry:
            # Entry exists - check if user can edit
            if user_role == 'Canteen':
                # Canteen staff cannot edit existing entries
                return jsonify({"error": "Canteen staff cannot edit existing entries"}), 403
            elif user_role == 'Admin':
                # Admin can edit. If 0, delete it.
                if amount == 0:
                    mongo.db.canteen_sales.delete_one({'_id': existing_entry['_id']})
                    return jsonify({"message": "Entry removed as amount was 0", "id": str(existing_entry['_id'])}), 200
                else:
                    mongo.db.canteen_sales.update_one(
                        {'_id': existing_entry['_id']},
                        {'$set': {
                            'amount': amount,
                            'edited_by': username,
                            'edited_at': datetime.now()
                        }}
                    )
                    return jsonify({"message": "Entry updated", "id": str(existing_entry['_id'])}), 200
        else:
            # New entry - both Admin and Canteen can add
            new_entry = {
                'patient_id': patient_id,
                'date': entry_date,
                'amount': amount,
                'entry_type': entry_type,
                'item': data.get('item', ''),  # Optional item description
                'recorded_by': username,
                'created_at': datetime.now()
            }
            result = mongo.db.canteen_sales.insert_one(new_entry)
            return jsonify({"message": "Entry recorded", "id": str(result.inserted_id)}), 201
            
    except ValueError as ve:
        return jsonify({"error": f"Invalid data format: {str(ve)}"}), 400
    except Exception as e:
        print(f"Daily Entry Error: {e}")
        return jsonify({"error": str(e)}), 500

# --- EXPENSES APIs ---

@app.route('/api/expenses', methods=['GET'])
@login_required
def list_expenses():
    if not check_db():
        return jsonify({"error": "Database error"}), 500
    try:
        cursor = mongo.db.expenses.find().sort('date', -1)
        expenses = []
        for e in cursor:
            expenses.append({
                'id': str(e.get('_id')),
                'type': e.get('type', 'outgoing'),
                'amount': e.get('amount', 0),
                'category': e.get('category', ''),
                'note': e.get('note', ''),
                'date': e.get('date').isoformat() if e.get('date') else '',
                'recorded_by': e.get('recorded_by', ''),
                'auto': False
            })

        # Automated income entries (not stored, just surfaced)
        try:
            # Monthly fees sum (all patients)
            patients = mongo.db.patients.find()
            total_fees = 0
            for p in patients:
                try:
                    total_fees += int(str(p.get('monthlyFee', '0')).replace(',', ''))
                except ValueError:
                    pass

            # Canteen sales sum (all time or could be month? align with summary -> month)
            today = datetime.now()
            start_of_month = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            pipeline = [
                {'$match': {'date': {'$gte': start_of_month}}},
                {'$group': {'_id': None, 'total_sales': {'$sum': '$amount'}}}
            ]
            sales_result = list(mongo.db.canteen_sales.aggregate(pipeline))
            total_canteen = sales_result[0]['total_sales'] if sales_result else 0

            today_iso = datetime.now().date().isoformat()
            expenses.insert(0, {
                'id': 'auto-canteen',
                'type': 'incoming',
                'amount': total_canteen,
                'category': 'Canteen Sales (auto)',
                'note': 'Automatically calculated from canteen sales this month',
                'date': today_iso,
                'recorded_by': 'system',
                'auto': True
            })
            expenses.insert(0, {
                'id': 'auto-fees',
                'type': 'incoming',
                'amount': total_fees,
                'category': 'Monthly Fees (auto)',
                'note': 'Automatically calculated from patient monthly fees',
                'date': today_iso,
                'recorded_by': 'system',
                'auto': True
            })
        except Exception as e:
            print(f"Auto income calc error: {e}")

        return jsonify(expenses)
    except Exception as e:
        print(f"Expenses list error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/expenses', methods=['POST'])
@role_required(['Admin'])
def add_expense():
    if not check_db():
        return jsonify({"error": "Database error"}), 500
    data = clean_input_data(request.json or {})
    required = ['type', 'amount', 'category']
    if not all(k in data for k in required):
        return jsonify({"error": "Missing fields"}), 400
    try:
        amount = int(str(data.get('amount', 0)).replace(',', ''))
    except ValueError:
        return jsonify({"error": "Amount must be a number"}), 400

    expense = {
        'type': data.get('type', 'outgoing'),
        'amount': amount,
        'category': data.get('category', ''),
        'note': data.get('note', ''),
        'date': datetime.fromisoformat(data.get('date')) if data.get('date') else datetime.now(),
        'recorded_by': session.get('username', 'System'),
        'created_at': datetime.now()
    }
    try:
        result = mongo.db.expenses.insert_one(expense)
        return jsonify({"message": "Expense saved", "id": str(result.inserted_id)}), 201
    except Exception as e:
        print(f"Add expense error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/expenses/<id>', methods=['DELETE'])
@role_required(['Admin'])
def delete_expense(id):
    if not check_db():
        return jsonify({"error": "Database error"}), 500
    try:
        result = mongo.db.expenses.delete_one({'_id': ObjectId(id)})
        if result.deleted_count:
            return jsonify({"message": "Expense deleted"})
        return jsonify({"error": "Expense not found"}), 404
    except Exception as e:
        print(f"Delete expense error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/expenses/summary', methods=['GET'])
@login_required
def expenses_summary():
    if not check_db():
        return jsonify({"error": "Database error"}), 500

    today = datetime.now()
    start_of_month = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    try:
        pipeline = [
            {'$match': {'date': {'$gte': start_of_month}}},
            {'$group': {'_id': '$type', 'total': {'$sum': '$amount'}}}
        ]
        summary_data = list(mongo.db.expenses.aggregate(pipeline))
        incoming = 0
        outgoing = 0
        for item in summary_data:
            if item['_id'] == 'incoming':
                incoming = item['total']
            elif item['_id'] == 'outgoing':
                outgoing = item['total']

        # Note: We only count manual expenses here.
        # Patient fees and canteen are tracked separately in accounts.
        # This avoids double-counting since receivedAmount already captures actual payments.
        
        return jsonify({
            'incoming': incoming,  # Only manual recorded income
            'outgoing': outgoing,  # Only manual recorded expenses
            'net': incoming - outgoing
        })
    except Exception as e:
        print(f"Expenses summary error: {e}")
        return jsonify({"error": str(e)}), 500

# --- EXPORT ROUTE (No change, retained for functionality) ---

@app.route('/api/export', methods=['POST'])
@role_required(['Admin', 'Doctor', 'Psychologist'])
def export_patients():
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        req_data = request.get_json() or {}
        selected_fields = req_data.get('fields', 'all')
        current_user = session.get('user') or {}
        is_admin = current_user.get('role') == 'Admin'
        print(f"Export request from user: {current_user.get('username')}, is_admin: {is_admin}")
        
        cursor = mongo.db.patients.find()
        patients_list = list(cursor)
        print(f"Found {len(patients_list)} patients")
        
        if not patients_list:
            return jsonify({"error": "No patients found"}), 404

        # Prepare Data (Ensure new fields are included)
        export_data = []
        for p in patients_list:
            # Convert ObjectId to string
            patient_id = str(p.get('_id', '')) if '_id' in p else ''
            
            row = {
                'name': p.get('name', ''),
                'fatherName': p.get('fatherName', ''),
                'admissionDate': p.get('admissionDate', ''),
                'idNo': p.get('idNo', '') if is_admin else '',
                'age': p.get('age', ''),
                'cnic': p.get('cnic', '') if is_admin else '',
                'contactNo': p.get('contactNo', '') if is_admin else '',
                'address': p.get('address', '') if is_admin else '',
                'complaint': p.get('complaint', ''),
                'guardianName': p.get('guardianName', '') if is_admin else '',
                'relation': p.get('relation', '') if is_admin else '',
                'drugProblem': p.get('drugProblem', ''),
                'maritalStatus': p.get('maritalStatus', ''),
                'prevAdmissions': p.get('prevAdmissions', ''),
                'monthlyFee': p.get('monthlyFee', '') if is_admin else '',
                'monthlyAllowance': p.get('monthlyAllowance', '') if is_admin else '',
                'created_at': p.get('created_at', '')
            }
            export_data.append(row)

        print(f"Prepared {len(export_data)} rows for export")
        df = pd.DataFrame(export_data)
        print(f"Created DataFrame with columns: {list(df.columns)}")

        if isinstance(selected_fields, list) and len(selected_fields) > 0:
            valid_fields = [f for f in selected_fields if f in df.columns]
            if valid_fields:
                df = df[valid_fields]

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Patients')
            
            # Configure A4 page setup
            worksheet = writer.sheets['Patients']
            worksheet.page_setup.paperSize = 9  # A4
            worksheet.page_setup.orientation = 'landscape'
            worksheet.page_setup.fitToWidth = 1
            worksheet.page_setup.fitToHeight = 0
            worksheet.print_options.horizontalCentered = True
            worksheet.page_margins.left = 0.5
            worksheet.page_margins.right = 0.5
            worksheet.page_margins.top = 0.75
            worksheet.page_margins.bottom = 0.75
        
        output.seek(0)
        print("Excel file created successfully")
        
        return send_file(
            output, 
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='patients_export.xlsx'
        )
    except ImportError as ie:
        print(f"ImportError in export: {ie}")
        return jsonify({"error": "Missing 'openpyxl' library"}), 500
    except Exception as e:
        print(f"Error in export: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"{type(e).__name__}: {str(e)}"}), 500


# --- NEW ACCOUNTS ROUTE (ADMIN ONLY) ---

@app.route('/api/accounts/summary', methods=['GET'])
@role_required(['Admin'])
def get_accounts_summary():
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        from datetime import datetime
        
        # Get all patients - Added 'isDischarged' to projection
        patients = list(mongo.db.patients.find({}, {
            'name': 1, 'fatherName': 1, 'admissionDate': 1, 
            'monthlyFee': 1, 'address': 1, 'age': 1,
            'laundryStatus': 1, 'laundryAmount': 1, 'receivedAmount': 1,
            'isDischarged': 1
        }))
        
        # Get total canteen sales per patient
        pipeline = [
            {'$group': {'_id': '$patient_id', 'total_sales': {'$sum': '$amount'}}}
        ]
        sales_data = list(mongo.db.canteen_sales.aggregate(pipeline))
        sales_map = {str(s['_id']): s['total_sales'] for s in sales_data}

        summary = []
        for p in patients:
            pid = str(p['_id'])
            
            # Calculate days elapsed from admission date
            admission_date = p.get('admissionDate')
            days_elapsed = 0
            if admission_date:
                try:
                    if isinstance(admission_date, str):
                        admission_dt = datetime.fromisoformat(admission_date.replace('Z', '+00:00'))
                    else:
                        admission_dt = admission_date
                    days_diff = (datetime.now() - admission_dt).days
                    days_elapsed = max(0, days_diff)
                except:
                    days_elapsed = 0
            
            # Get fees and calculate prorated fees
            monthly_fee = p.get('monthlyFee', '0')
            calculated_fee = calculate_prorated_fee(monthly_fee, days_elapsed)
            
            summary.append({
                'id': pid,
                'name': p.get('name', ''),
                'fatherName': p.get('fatherName', ''),
                'age': p.get('age', ''),
                'area': p.get('address', ''), 
                'admissionDate': p.get('admissionDate', ''),
                'monthlyFee': monthly_fee,
                'calculatedFee': calculated_fee,  # NEW: Prorated fee
                'daysElapsed': days_elapsed,  # NEW: Days elapsed for reference
                'canteenTotal': sales_map.get(pid, 0),
                'laundryStatus': p.get('laundryStatus', False),
                'laundryAmount': p.get('laundryAmount', 0),
                'receivedAmount': p.get('receivedAmount', '0'),
                'isDischarged': p.get('isDischarged', False) # <--- NEW: Return discharge status
            })
        
        return jsonify(summary)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# --- CALL & MEETING TRACKING APIs ---

@app.route('/api/call_meeting_tracker', methods=['GET'])
@login_required
def get_call_meeting_data():
    """Get all call and meeting data for the month"""
    if not check_db(): return jsonify({"error": "Database error"}), 500
    
    try:
        today = datetime.now()
        year = int(request.args.get('year', today.year))
        month = int(request.args.get('month', today.month))
        
        # Fetch all records for the current month
        records_cursor = mongo.db.call_meeting_tracker.find({
            'year': year,
            'month': month
        }).sort('day', 1)
        
        records = []
        for r in records_cursor:
            r['_id'] = str(r['_id'])
            r['status'] = r.get('status', r.get('type', 'Tick'))
            records.append(r)
        
        return jsonify(records)
    except Exception as e:
        print(f"Call/Meeting Fetch Error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/call_meeting_tracker', methods=['POST'])
@role_required(['Admin'])
def add_call_meeting_entry():
    """Add or update a call/meeting entry"""
    if not check_db(): return jsonify({"error": "Database error"}), 500
    
    data = clean_input_data(request.json)
    if not all(k in data for k in ['name', 'day', 'month', 'year', 'date_of_admission']):
        return jsonify({"error": "Missing fields"}), 400
    
    status_value = data.get('status') or data.get('type') or 'Meeting'
    if status_value not in ['Meeting', 'Call']:
        return jsonify({"error": "Type must be Meeting or Call"}), 400
    
    try:
        entry = {
            'name': data['name'],
            'day': int(data['day']),
            'month': int(data['month']),
            'year': int(data['year']),
            'type': status_value,
            'status': status_value,
            'date_of_admission': data['date_of_admission'],
            'recorded_by': session.get('username', 'Admin'),
            'created_at': datetime.now()
        }
        
        # Check if entry already exists for this person on this day/month/year
        existing = mongo.db.call_meeting_tracker.find_one({
            'name': data['name'],
            'day': int(data['day']),
            'month': int(data['month']),
            'year': int(data['year'])
        })
        
        if existing:
            # Update existing entry
            mongo.db.call_meeting_tracker.update_one({'_id': existing['_id']}, {'$set': entry})
            return jsonify({"message": "Entry updated", "id": str(existing['_id'])}), 200
        else:
            # Create new entry
            result = mongo.db.call_meeting_tracker.insert_one(entry)
            return jsonify({"message": "Entry added", "id": str(result.inserted_id)}), 201
    except Exception as e:
        print(f"Call/Meeting Add Error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/call_meeting_tracker/<id>', methods=['DELETE'])
@role_required(['Admin'])
def delete_call_meeting_entry(id):
    """Delete a call/meeting entry"""
    if not check_db(): return jsonify({"error": "Database error"}), 500
    
    try:
        result = mongo.db.call_meeting_tracker.delete_one({'_id': ObjectId(id)})
        if result.deleted_count > 0:
            return jsonify({"message": "Entry deleted"}), 200
        else:
            return jsonify({"error": "Entry not found"}), 404
    except Exception as e:
        print(f"Call/Meeting Delete Error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/call_meeting_tracker/summary/<int:month>/<int:year>', methods=['GET'])
@login_required
def get_call_meeting_summary(month, year):
    """Get summary of calls and meetings for the month"""
    if not check_db(): return jsonify({"error": "Database error"}), 500
    
    try:
        # Get all records for the month
        records_cursor = mongo.db.call_meeting_tracker.find({
            'year': year,
            'month': month
        })
        
        # Count tick / cross
        tick_count = 0
        cross_count = 0
        by_person = {}
        
        for r in records_cursor:
            record_status = (r.get('status') or r.get('type') or 'Meeting')
            record_status = record_status.capitalize()
            is_meeting = record_status == 'Meeting'
            tick_count += 1 if is_meeting else 0
            cross_count += 0 if is_meeting else 1
            
            person = r.get('name', 'Unknown')
            if person not in by_person:
                by_person[person] = {'Meeting': 0, 'Call': 0}
            by_person[person]['Meeting'] = by_person[person].get('Meeting', 0) + (1 if is_meeting else 0)
            by_person[person]['Call'] = by_person[person].get('Call', 0) + (0 if is_meeting else 1)
        
        return jsonify({
            'totalMeetings': tick_count,
            'totalCalls': cross_count,
            'byPerson': by_person
        })
    except Exception as e:
        print(f"Call/Meeting Summary Error: {e}")
        return jsonify({"error": str(e)}), 500


# --- UTILITY BILLS ROUTES ---

@app.route('/api/utility_bills', methods=['GET'])
@role_required(['Admin'])
def get_utility_bills():
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        month = request.args.get('month', type=int)
        year = request.args.get('year', type=int)
        
        query = {}
        if month and year:
            search_pattern = f"{year}-{month:02d}-"
            query['due_date'] = {'$regex': f'^{search_pattern}'}
            
        cursor = mongo.db.utility_bills.find(query).sort('due_date', 1)
        bills = []
        for b in cursor:
            bills.append({
                'id': str(b['_id']),
                'type': b.get('type', 'Other'),
                'provider': b.get('provider', ''),
                'amount': b.get('amount', 0),
                'due_date': b.get('due_date'),
                'ref_no': b.get('ref_no', ''),
                'status': 'Unpaid'
            })
        return jsonify(bills)
    except Exception as e:
        print(f"Bills Fetch Error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/utility_bills', methods=['POST'])
@role_required(['Admin'])
def add_utility_bill():
    if not check_db(): return jsonify({"error": "Database error"}), 500
    data = clean_input_data(request.json)
    try:
        bill = {
            'type': data.get('type', 'Other'), # Electricity, Gas, etc.
            'provider': data.get('provider', ''),
            'amount': int(data.get('amount', 0)),
            'due_date': data.get('due_date'),
            'ref_no': data.get('ref_no', ''),
            'created_at': datetime.now()
        }
        result = mongo.db.utility_bills.insert_one(bill)
        return jsonify({"message": "Bill added", "id": str(result.inserted_id)}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/utility_bills/<id>', methods=['DELETE'])
@role_required(['Admin'])
def pay_utility_bill(id):
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        # OPTIONAL: When deleting/paying, record it as an expense automatically
        bill = mongo.db.utility_bills.find_one({'_id': ObjectId(id)})
        if bill:
            # Add to expenses
            mongo.db.expenses.insert_one({
                'type': 'outgoing',
                'amount': bill['amount'],
                'category': 'Utility Bill',
                'note': f"Paid bill for {bill.get('type')} (Ref: {bill.get('ref_no')})",
                'date': datetime.now(),
                'recorded_by': session.get('username', 'Admin'),
                'created_at': datetime.now()
            })
            
        # Remove from bills collection
        mongo.db.utility_bills.delete_one({'_id': ObjectId(id)})
        return jsonify({"message": "Bill paid and removed"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# --- TEAM / EMPLOYEE MANAGEMENT ROUTES ---
@app.route('/api/employees', methods=['GET'])
@role_required(['Admin'])
def get_employees():
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        month = request.args.get('month', type=int)
        year = request.args.get('year', type=int)

        # Sort by name alphabetically
        cursor = mongo.db.employees.find().sort('name', 1)
        employees = []
        for e in cursor:
            advance_value = e.get('advance', '')
            if month and year:
                adv_year = e.get('advance_year')
                adv_month = e.get('advance_month')
                # Month-aware overhead view: advance automatically resets to 0 in a new month.
                if adv_year == year and adv_month == month:
                    advance_value = e.get('advance', '')
                else:
                    advance_value = 0

            employees.append({
                'id': str(e['_id']),
                'name': e.get('name', ''),
                'designation': e.get('designation', ''),
                'pay': e.get('pay', ''),
                'advance': advance_value,
                'duty_timings': e.get('duty_timings', ''),
                'date_of_joining': e.get('date_of_joining', ''),
                'cnic': e.get('cnic', ''),
                'phone': e.get('phone', '')
            })
        return jsonify(employees)
    except Exception as e:
        print(f"Employee Fetch Error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/employees', methods=['POST'])
@role_required(['Admin'])
def add_employee():
    if not check_db(): return jsonify({"error": "Database error"}), 500
    data = clean_input_data(request.json)
    try:
        current = datetime.now()
        advance_month = int(data.get('month') or current.month)
        advance_year = int(data.get('year') or current.year)
        employee = {
            'name': data.get('name'),
            'designation': data.get('designation'),
            'pay': data.get('pay', ''),
            'advance': data.get('advance', ''),
            'duty_timings': data.get('duty_timings', ''),
            'date_of_joining': data.get('date_of_joining', ''),
            'cnic': data.get('cnic', ''),
            'phone': data.get('phone', ''),
            'advance_month': advance_month,
            'advance_year': advance_year,
            'created_at': current
        }
        result = mongo.db.employees.insert_one(employee)
        return jsonify({"message": "Employee added", "id": str(result.inserted_id)}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/employees/<id>', methods=['PUT'])
@role_required(['Admin'])
def update_employee(id):
    if not check_db(): return jsonify({"error": "Database error"}), 500
    data = clean_input_data(request.json)
    month = request.args.get('month', type=int)
    year = request.args.get('year', type=int)
    # Remove _id from data if present to avoid immutable field error
    if '_id' in data: del data['_id']
    if month and year and 'advance' in data:
        data['advance_month'] = month
        data['advance_year'] = year
    try:
        mongo.db.employees.update_one({'_id': ObjectId(id)}, {'$set': data})
        return jsonify({"message": "Employee updated"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/employees/<id>', methods=['DELETE'])
@role_required(['Admin'])
def delete_employee(id):
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        mongo.db.employees.delete_one({'_id': ObjectId(id)})
        return jsonify({"message": "Employee deleted"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def _object_id_or_none(raw_id):
    try:
        return ObjectId(str(raw_id))
    except Exception:
        return None


def _extract_payment_patient_name(note=''):
    note = str(note or '')
    if 'Partial payment from ' in note:
        return note.split('Partial payment from ')[1].split(' via ')[0].strip()
    if 'Payment from ' in note:
        return note.split('Payment from ')[1].split(' via ')[0].strip()
    if 'Initial Advance from ' in note:
        return note.split('Initial Advance from ')[1].split(' (')[0].strip()
    return 'Unknown'


def _build_payment_record_note(existing_note, patient_name, payment_method):
    patient_name = patient_name or 'Unknown'
    existing_note = str(existing_note or '')
    large_amount_suffix = ' (Large Amount)' if '(Large Amount)' in existing_note else ''

    if existing_note.startswith('Initial Advance from '):
        return f"Initial Advance from {patient_name} (Admission)"
    if existing_note.startswith('Partial payment from '):
        return f"Partial payment from {patient_name} via {payment_method}{large_amount_suffix}"
    return f"Payment from {patient_name} via {payment_method}{large_amount_suffix}"


def _adjust_patient_received_amount(patient_id, delta_amount):
    patient_oid = _object_id_or_none(patient_id)
    if not patient_oid:
        return

    patient = mongo.db.patients.find_one({'_id': patient_oid}, {'receivedAmount': 1})
    if not patient:
        return

    current_received = _safe_int_amount(patient.get('receivedAmount', 0))
    new_total = max(current_received + _safe_int_amount(delta_amount), 0)
    mongo.db.patients.update_one(
        {'_id': patient_oid},
        {'$set': {'receivedAmount': str(new_total)}}
    )


@app.route('/api/payment-records', methods=['GET'])
@role_required(['Admin'])
def get_payment_records():
    """Fetch all payment records (receipts) from expenses collection."""
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        # Fetch all incoming payments from Patient Fee category
        payments = list(mongo.db.expenses.find({
            'type': 'incoming',
            'category': 'Patient Fee'
        }).sort('date', -1))  # Most recent first
        
        # Get all unique patient IDs from payments to fetch names in bulk
        patient_ids = list(set(p.get('patient_id') for p in payments if p.get('patient_id')))
        patient_map = {}
        if patient_ids:
            valid_ids = []
            for pid in patient_ids:
                oid = _object_id_or_none(pid)
                if oid:
                    valid_ids.append(oid)
            patients = list(mongo.db.patients.find({'_id': {'$in': valid_ids}}, {'name': 1}))
            patient_map = {str(pat['_id']): pat.get('name', 'Unknown') for pat in patients}

        # Process and format the records
        records = []
        for p in payments:
            # Prefer patient_id lookup, then fallback to note parsing for legacy data
            pid = p.get('patient_id')
            patient_name = patient_map.get(str(pid)) if pid else None
            
            if not patient_name:
                patient_name = _extract_payment_patient_name(p.get('note', ''))
            
            records.append({
                '_id': str(p['_id']),
                'patient_id': str(pid) if pid else '',
                'patient_name': patient_name,
                'amount': p.get('amount', 0),
                'date': p.get('date').strftime('%Y-%m-%d') if p.get('date') else 'N/A',
                'payment_method': p.get('payment_method', 'Cash'),
                'recorded_by': p.get('recorded_by', 'Admin'),
                'screenshot': p.get('screenshot', ''),
                'note': p.get('note', '')
            })
        
        return jsonify(records)
    except Exception as e:
        print(f"Payment Records Error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/payment-records/<id>', methods=['PUT'])
@role_required(['Admin'])
def update_payment_record(id):
    """Update a patient payment record and keep the patient's received total in sync."""
    if not check_db(): return jsonify({"error": "Database error"}), 500

    payment_oid = _object_id_or_none(id)
    if not payment_oid:
        return jsonify({"error": "Invalid payment record id"}), 400

    data = clean_input_data(request.json or {})
    amount = _safe_int_amount(data.get('amount'))
    if amount <= 0:
        return jsonify({"error": "Amount must be greater than zero"}), 400

    raw_date = str(data.get('date', '')).strip()
    try:
        payment_date = datetime.fromisoformat(raw_date) if raw_date else None
    except ValueError:
        return jsonify({"error": "Invalid payment date"}), 400

    try:
        payment_doc = mongo.db.expenses.find_one({
            '_id': payment_oid,
            'type': 'incoming',
            'category': 'Patient Fee'
        })
        if not payment_doc:
            return jsonify({"error": "Payment record not found"}), 404

        payment_method = data.get('payment_method') or payment_doc.get('payment_method', 'Cash')
        is_online_payment = str(payment_method).lower().startswith('online')
        screenshot = data.get('screenshot', payment_doc.get('screenshot', '')) if is_online_payment else ''

        patient_id = payment_doc.get('patient_id')
        patient_name = _extract_payment_patient_name(payment_doc.get('note', ''))
        patient_oid = _object_id_or_none(patient_id)
        if patient_oid:
            patient = mongo.db.patients.find_one({'_id': patient_oid}, {'name': 1})
            if patient and patient.get('name'):
                patient_name = patient.get('name')

        old_amount = _safe_int_amount(payment_doc.get('amount'))
        amount_delta = amount - old_amount
        if amount_delta:
            _adjust_patient_received_amount(patient_id, amount_delta)

        mongo.db.expenses.update_one(
            {'_id': payment_oid},
            {
                '$set': {
                    'amount': amount,
                    'payment_method': payment_method,
                    'screenshot': screenshot,
                    'date': payment_date or payment_doc.get('date') or datetime.now(),
                    'note': _build_payment_record_note(payment_doc.get('note', ''), patient_name, payment_method),
                    'updated_at': datetime.now()
                }
            }
        )
        return jsonify({"message": "Payment record updated"})
    except Exception as e:
        print(f"Payment Record Update Error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/payment-records/<id>', methods=['DELETE'])
@role_required(['Admin'])
def delete_payment_record(id):
    """Delete a patient payment record and reverse its effect from the patient balance."""
    if not check_db(): return jsonify({"error": "Database error"}), 500

    payment_oid = _object_id_or_none(id)
    if not payment_oid:
        return jsonify({"error": "Invalid payment record id"}), 400

    try:
        payment_doc = mongo.db.expenses.find_one({
            '_id': payment_oid,
            'type': 'incoming',
            'category': 'Patient Fee'
        })
        if not payment_doc:
            return jsonify({"error": "Payment record not found"}), 404

        _adjust_patient_received_amount(
            payment_doc.get('patient_id'),
            -_safe_int_amount(payment_doc.get('amount'))
        )

        mongo.db.expenses.delete_one({'_id': payment_oid})
        return jsonify({"message": "Payment record deleted"})
    except Exception as e:
        print(f"Payment Record Delete Error: {e}")
        return jsonify({"error": str(e)}), 500


# ============================================================
#  OVERHEADS MANAGEMENT (Admin Only)
# ============================================================

@app.route('/api/finance/summary/<int:month>/<int:year>', methods=['GET'])
@role_required(['Admin'])
def get_finance_summary(month, year):
    """
    Get a complete financial summary for a given month/year.
    Includes expected incoming fees, salaries, utility bills,
    and daily overheads (kitchen, canteen, others).
    """
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        # Month boundaries for month/year-aware active patient filtering
        month_start = datetime(year, month, 1).date()
        if month == 12:
            month_end = datetime(year + 1, 1, 1).date()
        else:
            month_end = datetime(year, month + 1, 1).date()

        def _parse_date_only(raw_val):
            if not raw_val:
                return None
            if isinstance(raw_val, datetime):
                return raw_val.date()
            try:
                # Handles ISO datetime/date strings.
                return datetime.fromisoformat(str(raw_val).replace('Z', '+00:00')).date()
            except Exception:
                try:
                    # Fallback for non-ISO text that still starts with YYYY-MM-DD.
                    return datetime.strptime(str(raw_val)[:10], '%Y-%m-%d').date()
                except Exception:
                    return None

        # 0. Expected incoming from patients active in selected month
        # Card label: "Fees from Active Patients"
        patients = list(mongo.db.patients.find({}, {
            'monthlyFee': 1,
            'admissionDate': 1,
            'isDischarged': 1,
            'dischargeDate': 1
        }))

        total_expected_incoming = 0
        for patient in patients:
            admission_date = _parse_date_only(patient.get('admissionDate'))
            discharge_date = _parse_date_only(patient.get('dischargeDate')) if patient.get('isDischarged') else None

            # Include patient only if they overlap selected month.
            if admission_date and admission_date >= month_end:
                continue
            if discharge_date and discharge_date < month_start:
                continue

            total_expected_incoming += _safe_int_amount(patient.get('monthlyFee', 0))

        # 1. Salaries Total
        employees = list(mongo.db.employees.find())
        total_salaries = 0
        for emp in employees:
            try:
                pay = int(str(emp.get('pay', '0')).replace(',', ''))
                total_salaries += pay
            except: pass

        # 2. Utility Bills for this month
        # We look for bills with a due_date in the format YYYY-MM-DD that matches
        search_pattern = f"{year}-{month:02d}-"
        bills = list(mongo.db.utility_bills.find({'due_date': {'$regex': f'^{search_pattern}'}}))
        total_bills = sum(b.get('amount', 0) for b in bills)

        # 3. Daily Overheads (Expenses and Income)
        overheads = list(mongo.db.overheads.find({
            'month': month,
            'year': year
        }))
        
        total_kitchen = 0
        total_others = 0
        total_pay_advance = 0
        
        for entry in overheads:
            total_kitchen += entry.get('kitchen', 0)
            total_others += entry.get('others', 0)
            total_pay_advance += entry.get('pay_advance', 0)

        # 4. Canteen Auto (Calculated from canteen_sales for the month)
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)
            
        canteen_agg = mongo.db.canteen_sales.aggregate([
            {'$match': {'date': {'$gte': start_date, '$lt': end_date}}},
            {'$group': {'_id': None, 'total': {'$sum': '$amount'}}}
        ])
        canteen_res = list(canteen_agg)
        total_canteen_auto = canteen_res[0]['total'] if canteen_res else 0

        # Grand Total Overheads
        total_estimated_overheads = total_salaries + total_bills + total_kitchen + total_others + total_pay_advance + total_canteen_auto

        return jsonify({
            'month': month,
            'year': year,
            'totalSalaries': total_salaries,
            'totalUtilityBills': total_bills,
            'totalKitchen': total_kitchen,
            'totalCanteenAuto': total_canteen_auto,
            'totalOthers': total_others,
            'totalPayAdvance': total_pay_advance,
            'totalEstimatedOverheads': total_estimated_overheads,
            'totalIncome': total_expected_incoming,
            'profit': total_expected_incoming - total_estimated_overheads
        })
    except Exception as e:
        print(f"Finance Summary Error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/inventory/stats/<int:month>/<int:year>', methods=['GET'])
@login_required
def get_inventory_stats(month, year):
    """
    Get inventory stats for a given month/year.
    Returns: new patient admissions, discharges, total canteen sales.
    Accessible to all logged-in users.
    """
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        month_start = datetime(year, month, 1)
        if month == 12:
            month_end = datetime(year + 1, 1, 1)
        else:
            month_end = datetime(year, month + 1, 1)

        def parse_date(raw):
            if not raw: return None
            if isinstance(raw, datetime): return raw.replace(tzinfo=None)
            try:
                return datetime.fromisoformat(str(raw).replace('Z', '+00:00')).replace(tzinfo=None)
            except Exception:
                try:
                    return datetime.strptime(str(raw)[:10], '%Y-%m-%d')
                except Exception:
                    return None

        all_patients = list(mongo.db.patients.find({}, {
            'admissionDate': 1, 'isDischarged': 1, 'dischargeDate': 1
        }))

        new_count = 0
        discharged_count = 0
        for p in all_patients:
            admit = parse_date(p.get('admissionDate'))
            if admit and month_start <= admit < month_end:
                new_count += 1
            if p.get('isDischarged'):
                disch = parse_date(p.get('dischargeDate'))
                if disch and month_start <= disch < month_end:
                    discharged_count += 1

        canteen_result = list(mongo.db.canteen_sales.aggregate([
            {
                '$match': {
                    'date': {'$gte': month_start, '$lt': month_end},
                    '$or': [
                        {'entry_type': {'$exists': False}},
                        {'entry_type': {'$ne': 'other'}}
                    ]
                }
            },
            {'$group': {'_id': None, 'total': {'$sum': '$amount'}}}
        ]))
        total_canteen = canteen_result[0]['total'] if canteen_result else 0

        return jsonify({
            'new_patients': new_count,
            'discharged': discharged_count,
            'total_canteen_sales': total_canteen
        })
    except Exception as e:
        print(f"Inventory Stats Error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/overheads/<int:month>/<int:year>', methods=['GET'])
@role_required(['Admin'])
def get_overheads(month, year):
    """
    Fetch overhead entries for a given month/year.
    Also aggregates daily canteen totals from canteen_sales.
    """
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        # Fetch stored overhead entries for this month
        overheads = list(mongo.db.overheads.find({
            'month': month,
            'year': year
        }))
        
        # Convert to dict keyed by date
        overhead_map = {}
        for entry in overheads:
            date_key = entry.get('date')
            if date_key:
                overhead_map[date_key] = {
                    '_id': str(entry['_id']),
                    'date': date_key,
                    'kitchen': entry.get('kitchen', 0),
                    'canteen_auto': entry.get('canteen_auto', 0),
                    'others': entry.get('others', 0),
                    'pay_advance': entry.get('pay_advance', 0),
                    'employee_names': entry.get('employee_names', ''),
                    'income': entry.get('income', 0),
                    'total_expense': entry.get('total_expense', 0)
                }
        
        # Aggregate daily canteen sales totals
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)
        
        canteen_aggregation = mongo.db.canteen_sales.aggregate([
            {
                '$match': {
                    'date': {
                        '$gte': start_date,
                        '$lt': end_date
                    },
                    '$or': [
                        {'entry_type': {'$exists': False}},
                        {'entry_type': {'$ne': 'other'}}
                    ]
                }
            },
            {
                '$group': {
                    '_id': {
                        '$dateToString': {
                            'format': '%Y-%m-%d',
                            'date': '$date'
                        }
                    },
                    'total': {'$sum': '$amount'}
                }
            }
        ])
        
        canteen_daily = {item['_id']: item['total'] for item in canteen_aggregation}
        
        # Calculate days in month
        if month == 12:
            next_month = datetime(year + 1, 1, 1)
        else:
            next_month = datetime(year, month + 1, 1)
        days_in_month = (next_month - datetime(year, month, 1)).days
        
        return jsonify({
            'overheads': overhead_map,
            'canteen_daily': canteen_daily,
            'days_in_month': days_in_month
        })
    except Exception as e:
        print(f"Get Overheads Error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/overheads/annual/<int:year>', methods=['GET'])
@role_required(['Admin'])
def get_overheads_annual(year):
    """
    Aggregate total income, expense, and profit for a full year,
    including canteen sales from canteen_sales collection.
    """
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        # 1. Aggregate canteen sales day-by-day for the year
        canteen_daily_aggr = mongo.db.canteen_sales.aggregate([
            {
                '$match': {
                    'date': {'$gte': start_date, '$lt': end_date},
                    '$or': [
                        {'entry_type': {'$exists': False}},
                        {'entry_type': {'$ne': 'other'}}
                    ]
                }
            },
            {
                '$group': {
                    '_id': {'$dateToString': {'format': '%Y-%m-%d', 'date': '$date'}},
                    'total': {'$sum': '$amount'}
                }
            }
        ])
        canteen_daily_map = {item['_id']: item['total'] for item in canteen_daily_aggr}
        
        # 2. Fetch all overhead entries for the year
        entries = list(mongo.db.overheads.find({'year': year}))
        overhead_map = {e.get('date'): e for e in entries if e.get('date')}

        total_income = 0.0
        total_expense = 0.0
        total_canteen = 0.0

        # 3. Calculate totals day-by-day (or at least for all dates present in either map)
        all_dates = set(canteen_daily_map.keys()) | set(overhead_map.keys())
        
        for d_str in all_dates:
            entry = overhead_map.get(d_str, {})
            # Use stored canteen_auto (override) if exists, else use raw sales
            day_canteen = entry.get('canteen_auto') if entry.get('canteen_auto') is not None else canteen_daily_map.get(d_str, 0)
            
            day_kitchen = float(entry.get('kitchen', 0))
            day_others = float(entry.get('others', 0))
            day_pay_advance = float(entry.get('pay_advance', 0))
            day_income = float(entry.get('income', 0))
            
            day_expense = day_kitchen + day_canteen + day_others + day_pay_advance
            
            total_income += day_income
            total_expense += day_expense
            total_canteen += day_canteen

        return jsonify({
            'year': year,
            'total_income': total_income,
            'total_expense': total_expense,
            'total_canteen': total_canteen,
            'profit': total_income - total_expense
        })
    except Exception as e:
        print(f"Get Annual Overheads Error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/overheads/entry', methods=['POST'])
@role_required(['Admin'])
def save_overhead_entry():
    """
    Save or update a single overhead entry for a specific date.
    """
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        data = request.get_json()
        date = data.get('date')  # Format: YYYY-MM-DD
        month = data.get('month')
        year = data.get('year')
        
        # Parse values
        kitchen = float(data.get('kitchen', 0))
        others = float(data.get('others', 0))
        pay_advance = float(data.get('pay_advance', 0))
        income = float(data.get('income', 0))
        employee_names = data.get('employee_names', '')
        canteen_auto = float(data.get('canteen_auto', 0))
        
        # Calculate total expense
        total_expense = kitchen + canteen_auto + others + pay_advance
        
        entry = {
            'date': date,
            'month': month,
            'year': year,
            'kitchen': kitchen,
            'canteen_auto': canteen_auto,
            'others': others,
            'pay_advance': pay_advance,
            'employee_names': employee_names,
            'income': income,
            'total_expense': total_expense,
            'last_updated': datetime.now()
        }
        
        # Upsert: update if exists, insert if not
        mongo.db.overheads.update_one(
            {'date': date, 'month': month, 'year': year},
            {'$set': entry},
            upsert=True
        )
        
        return jsonify({"message": "Entry saved", "entry": entry})
    except Exception as e:
        print(f"Save Overhead Entry Error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/overheads/canteen-sync/<int:month>/<int:year>', methods=['GET'])
@role_required(['Admin'])
def sync_overheads_canteen(month, year):
    """
    Get updated daily canteen totals for the month.
    Used for real-time sync when canteen sales are added.
    """
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)
        
        canteen_aggregation = mongo.db.canteen_sales.aggregate([
            {
                '$match': {
                    'date': {
                        '$gte': start_date,
                        '$lt': end_date
                    }
                }
            },
            {
                '$group': {
                    '_id': {
                        '$dateToString': {
                            'format': '%Y-%m-%d',
                            'date': '$date'
                        }
                    },
                    'total': {'$sum': '$amount'}
                }
            }
        ])
        
        canteen_daily = {item['_id']: item['total'] for item in canteen_aggregation}
        
        return jsonify({'canteen_daily': canteen_daily})
    except Exception as e:
        print(f"Sync Overheads Canteen Error: {e}")
        return jsonify({"error": str(e)}), 500


def _month_start_n_months_ago(months_ago: int) -> datetime:
    today = datetime.now()
    # months_ago = 0 => current month start; 5 => 5 months back
    target_month = today.month - months_ago
    target_year = today.year
    while target_month <= 0:
        target_month += 12
        target_year -= 1
    return datetime(target_year, target_month, 1)


@app.route('/api/payment-records/export', methods=['GET'])
@role_required(['Admin'])
def export_payment_records():
    """Export payment records to Excel for a given range.
    range=current (default) or six_months.
    """
    if not check_db(): return jsonify({"error": "Database error"}), 500

    range_key = request.args.get('range', 'current')

    today = datetime.now()
    if range_key == 'six_months':
        start_date = _month_start_n_months_ago(5)  # includes current month (6 total)
    else:
        start_date = datetime(today.year, today.month, 1)

    # end date = first day of next month
    if today.month == 12:
        end_date = datetime(today.year + 1, 1, 1)
    else:
        end_date = datetime(today.year, today.month + 1, 1)

    try:
        payments = list(mongo.db.expenses.find({
            'type': 'incoming',
            'category': 'Patient Fee',
            'date': {'$gte': start_date, '$lt': end_date}
        }).sort('date', 1))

        rows = []

        def to_date(dt_val):
            if not dt_val:
                return ''
            if isinstance(dt_val, datetime):
                return dt_val
            try:
                return datetime.fromisoformat(str(dt_val))
            except Exception:
                return None

        # Get all unique patient IDs from payments to fetch names in bulk
        patient_ids = list(set(p.get('patient_id') for p in payments if p.get('patient_id')))
        patient_map = {}
        if patient_ids:
            valid_ids = []
            for pid in patient_ids:
                oid = _object_id_or_none(pid)
                if oid:
                    valid_ids.append(oid)
            patients = list(mongo.db.patients.find({'_id': {'$in': valid_ids}}, {'name': 1}))
            patient_map = {str(pat['_id']): pat.get('name', 'Unknown') for pat in patients}

        for p in payments:
            # Prefer patient_id lookup, then fallback to note parsing
            pid = p.get('patient_id')
            note = p.get('note', '')
            patient_name = patient_map.get(str(pid)) if pid else None
            
            if not patient_name:
                patient_name = _extract_payment_patient_name(note)

            dt = to_date(p.get('date'))
            rows.append({
                'Patient Name': patient_name,
                'Amount (PKR)': p.get('amount', 0),
                'Date': dt.strftime('%Y-%m-%d') if dt else '',
                'Payment Mode': p.get('payment_method', 'Cash'),
                'Recorded By': p.get('recorded_by', 'Admin'),
                'Note': note
            })

        df = pd.DataFrame(rows)
        if df.empty:
            df = pd.DataFrame([{'Message': 'No payment records for selected range'}])

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Payment Records')
            
            # Configure A4 page setup
            worksheet = writer.sheets['Payment Records']
            worksheet.page_setup.paperSize = 9  # A4
            worksheet.page_setup.orientation = 'portrait'
            worksheet.page_setup.fitToWidth = 1
            worksheet.page_setup.fitToHeight = 0
            worksheet.print_options.horizontalCentered = True
            worksheet.page_margins.left = 0.75
            worksheet.page_margins.right = 0.75
            worksheet.page_margins.top = 0.75
            worksheet.page_margins.bottom = 0.75
        
        output.seek(0)

        filename = f"payment_records_{'six_months' if range_key == 'six_months' else 'current_month'}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        print(f"Payment Records Export Error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/patients/<id>/payment', methods=['POST'])
@role_required(['Admin'])
def add_patient_payment(id):
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        data = clean_input_data(request.json or {})
        amount_paid = _safe_int_amount(data.get('amount'))
        if amount_paid <= 0:
            return jsonify({"error": "Amount must be greater than zero"}), 400
        payment_method = data.get('payment_method', 'Cash') # Cash or Online
        screenshot = data.get('screenshot', '') if str(payment_method).lower().startswith('online') else ''
        raw_payment_date = str(data.get('payment_date', '')).strip()
        try:
            payment_date = datetime.fromisoformat(raw_payment_date) if raw_payment_date else datetime.now()
        except ValueError:
            return jsonify({"error": "Invalid payment date"}), 400
        
        patient = mongo.db.patients.find_one({'_id': ObjectId(id)})
        if not patient:
            return jsonify({"error": "Patient not found"}), 404

        # 1. Parse existing received amount
        current_received_str = str(patient.get('receivedAmount', '0')).replace(',', '')
        try:
            current_received = int(current_received_str)
        except ValueError:
            current_received = 0

        # 2. Add new payment
        new_total = current_received + amount_paid

        # 3. Update Patient Record
        mongo.db.patients.update_one(
            {'_id': ObjectId(id)}, 
            {'$set': {'receivedAmount': str(new_total)}}
        )

        # 4. Log as an Incoming Expense automatically
        is_overpayment = False
        try:
            # Simple check: if amount is very large, mark it. 
            # In a full system, we'd calculate prorated fee here too, 
            # but frontend already warned the user.
            if amount_paid > 100000: # Example threshold for suspicious amount
                is_overpayment = True
        except: pass

        expense_note = f"Payment from {patient.get('name')} via {payment_method}"
        if is_overpayment: expense_note += " (Large Amount)"
        
        mongo.db.expenses.insert_one({
            'type': 'incoming',
            'amount': amount_paid,
            'category': 'Patient Fee',
            'note': expense_note,
            'payment_method': payment_method,
            'patient_id': str(id),
            'screenshot': screenshot,
            'date': payment_date,
            'recorded_by': session.get('username', 'Admin'),
            'auto': True
        })

        return jsonify({"message": "Payment recorded successfully", "new_total": new_total})
    except Exception as e:
        print(f"Payment Error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/patients/<id>/discharge-bill', methods=['GET'])
@role_required(['Admin', 'Doctor'])
def generate_discharge_bill(id):
    """Generate a discharge bill for a patient - formatted to fit on one A4 page"""
    if not check_db(): return jsonify({"error": "Database error"}), 500
    
    try:
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        
        # Fetch patient data
        patient = mongo.db.patients.find_one({'_id': ObjectId(id)})
        if not patient:
            return jsonify({"error": "Patient not found"}), 404
        
        # Calculate days elapsed from admission date
        admission_date = patient.get('admissionDate')
        days_elapsed = 0
        if admission_date:
            try:
                if isinstance(admission_date, str):
                    admission_dt = datetime.fromisoformat(admission_date.replace('Z', '+00:00'))
                else:
                    admission_dt = admission_date
                days_diff = (datetime.now() - admission_dt).days
                days_elapsed = max(0, days_diff)
            except:
                days_elapsed = 0
        
        # Calculate canteen sales total for this patient
        pipeline = [
            {'$match': {'patient_id': ObjectId(id)}},
            {'$group': {'_id': None, 'total_sales': {'$sum': '$amount'}}}
        ]
        canteen_result = list(mongo.db.canteen_sales.aggregate(pipeline))
        canteen_total = canteen_result[0]['total_sales'] if canteen_result else 0
        
        # Parse financial data and calculate prorated/per-day fee
        monthly_fee_raw = patient.get('monthlyFee', '0')
        monthly_fee = calculate_prorated_fee(monthly_fee_raw, days_elapsed)
        
        laundry_amount = patient.get('laundryAmount', 0) if patient.get('laundryStatus', False) else 0
        received_amount = int(str(patient.get('receivedAmount', '0')).replace(',', '') or '0')
        
        # Calculate totals
        total_charges = monthly_fee + canteen_total + laundry_amount
        balance_due = total_charges - received_amount
        
        # Create discharge bill data
        bill_data = {
            'Patient Name': patient.get('name', ''),
            'Father Name': patient.get('fatherName', ''),
            'CNIC': patient.get('cnic', ''),
            'Admission Date': patient.get('admissionDate', ''),
            'Discharge Date': patient.get('dischargeDate', '') or datetime.now().strftime('%Y-%m-%d'),
            'Days Stayed': days_elapsed,
            'Monthly Fee': monthly_fee,
            'Canteen Charges': canteen_total,
            'Laundry Charges': laundry_amount,
            'Total Charges': total_charges,
            'Amount Paid': received_amount,
            'Balance Due': balance_due
        }
        
        # Create Excel workbook
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Create a DataFrame for the bill
            df = pd.DataFrame([bill_data])
            df.to_excel(writer, index=False, sheet_name='Discharge Bill')
            
            worksheet = writer.sheets['Discharge Bill']
            
            # Configure A4 page setup - Portrait, fit to one page
            worksheet.page_setup.paperSize = 9  # A4
            worksheet.page_setup.orientation = 'portrait'
            worksheet.page_setup.fitToPage = True
            worksheet.page_setup.fitToWidth = 1
            worksheet.page_setup.fitToHeight = 1  # Force to fit on 1 page height
            worksheet.page_setup.scale = None  # Allow auto-scaling
            
            # Set margins to maximize space
            worksheet.page_margins.left = 0.5
            worksheet.page_margins.right = 0.5
            worksheet.page_margins.top = 0.5
            worksheet.page_margins.bottom = 0.5
            worksheet.page_margins.header = 0.3
            worksheet.page_margins.footer = 0.3
            
            # Center horizontally on page
            worksheet.print_options.horizontalCentered = True
            
            # Styling
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            
            # Apply styles to header row
            for cell in worksheet[1]:
                cell.font = Font(bold=True, size=10, color='FFFFFF')
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            # Apply styles to data rows
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.font = Font(size=10)
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    cell.border = thin_border
                    # Highlight financial totals
                    if 'Total' in str(worksheet.cell(1, cell.column).value) or 'Balance' in str(worksheet.cell(1, cell.column).value):
                        cell.fill = total_fill
                        cell.font = Font(size=10, bold=True)
            
            # Auto-adjust column widths (but keep them reasonable for A4)
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 25)  # Cap at 25 to fit on A4
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Set row height
            worksheet.row_dimensions[1].height = 30  # Header
            for row in range(2, worksheet.max_row + 1):
                worksheet.row_dimensions[row].height = 20
        
        output.seek(0)
        
        filename = f"discharge_bill_{patient.get('name', 'patient').replace(' ', '_')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"Discharge Bill Error: {e}")
        return jsonify({"error": str(e)}), 500


# --- DAILY REPORT APIS ---

@app.route('/api/reports', methods=['GET'])
@role_required(['Admin', 'General Staff', 'Doctor'])
def get_daily_report():
    if not check_db(): return jsonify({"error": "Database error"}), 500
    
    date_str = request.args.get('date')
    if not date_str:
        return jsonify({"error": "Date required"}), 400
        
    try:
        # Fetch all report entries for this specific date
        reports = list(mongo.db.daily_reports.find({'date': date_str}))
        
        # Convert ObjectId to string
        for r in reports:
            r['_id'] = str(r['_id'])
            r['patient_id'] = str(r['patient_id'])
            
        return jsonify(reports)
    except Exception as e:
        print(f"Report Fetch Error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/reports/update', methods=['POST'])
@role_required(['Admin', 'General Staff', 'Doctor'])
def update_daily_report():
    if not check_db(): return jsonify({"error": "Database error"}), 500
    
    data = clean_input_data(request.json)
    # Expected: { date, patient_id, time_slot, status }
    # status enum: 'done', 'not_done', 'complaint', ''
    
    try:
        query = {
            'date': data['date'],
            'patient_id': ObjectId(data['patient_id'])
        }
        
        # Upsert: Update if exists, Insert if not
        update = {
            '$set': {
                f"schedule.{data['time_slot']}": data['status'],
                'updated_at': datetime.now(),
                'updated_by': session.get('username', 'System')
            }
        }
        
        mongo.db.daily_reports.update_one(query, update, upsert=True)
        return jsonify({"message": "Status updated"}), 200
        
    except Exception as e:
        print(f"Report Update Error: {e}")
        return jsonify({"error": str(e)}), 500

# --- REPORT CONFIGURATION API ---
    
@app.route('/api/reports/config', methods=['GET'])
@login_required
def get_report_config():
    if not check_db(): return jsonify({})
    # Return saved config or empty (frontend handles defaults)
    config = mongo.db.report_config.find_one({'_id': 'main_config'})
    if config:
        return jsonify(config)
    return jsonify({})

@app.route('/api/reports/config', methods=['POST'])
@role_required(['Admin'])
def save_report_config():
    if not check_db(): return jsonify({"error": "Database error"}), 500
    data = clean_input_data(request.json)
    try:
        # Save day_columns and night_columns
        mongo.db.report_config.update_one(
            {'_id': 'main_config'},
            {'$set': {
                'day_columns': data.get('day_columns'),
                'night_columns': data.get('night_columns')
            }},
            upsert=True
        )
        return jsonify({"message": "Layout saved"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# --- PSYCHOLOGIST SESSIONS ---

def _parse_iso_date(date_str):
    try:
        return datetime.fromisoformat(date_str)
    except Exception:
        return None


@app.route('/api/psych-sessions', methods=['GET'])
@login_required
def list_psych_sessions():
    if not check_db():
        return jsonify({"error": "Database error"}), 500

    role = session.get('role')
    user_id = session.get('user_id')

    start_str = request.args.get('start')
    end_str = request.args.get('end')
    psychologist_id = request.args.get('psychologistId')

    start_date = _parse_iso_date(start_str) if start_str else None
    end_date = _parse_iso_date(end_str) if end_str else None

    if end_date:
        # make end exclusive by moving to next day start
        end_date = end_date + timedelta(days=1)

    query = {}
    if start_date and end_date:
        query['date'] = {'$gte': start_date, '$lt': end_date}
    elif start_date:
        query['date'] = {'$gte': start_date}

    if role == 'Psychologist':
        query['psychologist_id'] = user_id
    elif psychologist_id:
        query['psychologist_id'] = psychologist_id

    try:
        sessions_cursor = mongo.db.psych_sessions.find(query).sort('date', 1)
        sessions = list(sessions_cursor)

        # collect ids for enrichment
        patient_ids = set()
        psych_ids = set()
        for s in sessions:
            for pid in s.get('patient_ids', []):
                patient_ids.add(pid)
            if s.get('psychologist_id'):
                psych_ids.add(s.get('psychologist_id'))

        patient_map = {}
        if patient_ids:
            patients = mongo.db.patients.find({"_id": {"$in": [ObjectId(pid) for pid in patient_ids if ObjectId.is_valid(pid)]}})
            for p in patients:
                patient_map[str(p['_id'])] = p.get('name', 'Unknown')

        psych_map = {}
        if psych_ids:
            users = mongo.db.users.find({"_id": {"$in": [ObjectId(pid) for pid in psych_ids if ObjectId.is_valid(pid)]}})
            for u in users:
                psych_map[str(u['_id'])] = u.get('name', u.get('username', 'Psych'))

        result = []
        for s in sessions:
            result.append({
                '_id': str(s['_id']),
                'psychologist_id': s.get('psychologist_id'),
                'psychologist_name': psych_map.get(s.get('psychologist_id', ''), s.get('psychologist_id', '')),
                'date': s.get('date').strftime('%Y-%m-%d') if s.get('date') else '',
                'time_slot': s.get('time_slot', ''),
                'patient_ids': s.get('patient_ids', []),
                'patient_names': [patient_map.get(pid, 'Unknown') for pid in s.get('patient_ids', [])],
                'title': s.get('title', ''),
                'note': s.get('note', ''),
                'note_detail': s.get('note_detail'),
                'note_author': s.get('note_author', ''),
                'note_at': s.get('note_at').isoformat() if s.get('note_at') else None
            })

        return jsonify(result)
    except Exception as e:
        print(f"Psych sessions fetch error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/psych-sessions', methods=['POST'])
@role_required(['Admin'])
def create_psych_session():
    if not check_db():
        return jsonify({"error": "Database error"}), 500

    data = clean_input_data(request.json)
    date_str = data.get('date')
    time_slot = data.get('time_slot', '')
    psychologist_id = data.get('psychologist_id')
    patient_ids = data.get('patient_ids', []) or []
    title = data.get('title', '')

    if not (date_str and psychologist_id and patient_ids):
        return jsonify({"error": "Missing fields"}), 400

    date_val = _parse_iso_date(date_str)
    if not date_val:
        return jsonify({"error": "Invalid date"}), 400

    # normalize to date-only at midnight
    date_val = date_val.replace(hour=0, minute=0, second=0, microsecond=0)

    try:
        doc = {
            'psychologist_id': psychologist_id,
            'date': date_val,
            'time_slot': time_slot,
            'patient_ids': patient_ids,
            'title': title,
            'created_by': session.get('username'),
            'created_at': datetime.now()
        }

        res = mongo.db.psych_sessions.insert_one(doc)
        return jsonify({"message": "Session created", "id": str(res.inserted_id)})
    except Exception as e:
        print(f"Psych session create error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/psych-sessions/<session_id>/note', methods=['POST'])
@role_required(['Admin', 'Psychologist'])
def add_psych_session_note(session_id):
    if not check_db():
        return jsonify({"error": "Database error"}), 500

    data = clean_input_data(request.json)
    note_text = data.get('note', '').strip()
    note_issue = data.get('issue', '').strip()
    note_intervention = data.get('intervention', '').strip()
    note_response = data.get('response', '').strip()

    # Require the structured fields; keep legacy fallback if only note provided
    if not (note_issue and note_intervention and note_response):
        if not note_text:
            return jsonify({"error": "Issue, intervention, and response are required"}), 400
    else:
        # Compose a legacy note string for compatibility
        note_text = f"Issue: {note_issue}\nIntervention: {note_intervention}\nResponse: {note_response}"

    try:
        session_doc = mongo.db.psych_sessions.find_one({'_id': ObjectId(session_id)})
        if not session_doc:
            return jsonify({"error": "Session not found"}), 404

        if session_doc.get('note'):
            return jsonify({"error": "Note already saved"}), 409

        mongo.db.psych_sessions.update_one(
            {'_id': ObjectId(session_id)},
            {'$set': {
                'note': note_text,
                'note_detail': {
                    'issue': note_issue,
                    'intervention': note_intervention,
                    'response': note_response
                } if note_issue and note_intervention and note_response else None,
                'note_author': session.get('username'),
                'note_at': datetime.now()
            }}
        )

        return jsonify({"message": "Note saved"})
    except Exception as e:
        print(f"Psych session note error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/attendance')
def get_attendance():
    if not check_db(): return jsonify({"error": "Database error"}), 500
    year = int(request.args.get('year'))
    month = int(request.args.get('month'))

    records = mongo.db.attendance.find({
        "year": year,
        "month": month
    })

    result = {}
    for rec in records:
        emp_id = str(rec["employee_id"])
        result[emp_id] = rec.get("days", {})

    return jsonify(result)

@app.route('/api/attendance', methods=['POST'])
def save_attendance():
    if not check_db(): return jsonify({"error": "Database error"}), 500
    data = request.json

    employee_id = data['empId']
    day = str(data['day'])
    year = int(data['year'])
    month = int(data['month'])
    mark = data['mark']  # 'P', 'A', or ''

    query = {
        "employee_id": employee_id,
        "year": year,
        "month": month
    }

    if mark == '':
        mongo.db.attendance.update_one(
            query,
            { "$unset": { f"days.{day}": "" } },
            upsert=True
        )
    else:
        mongo.db.attendance.update_one(
            query,
            { "$set": { f"days.{day}": mark } },
            upsert=True
        )

    return jsonify(success=True)

# --- EMERGENCY DASHBOARD APIs ---
@app.route('/api/emergency', methods=['GET'])
@login_required
def get_emergency_alerts():
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        # Only return active alerts (where status != 'resolved')
        alerts = list(mongo.db.emergency_alerts.find({'status': {'$ne': 'resolved'}}).sort('created_at', -1))
        for a in alerts:
            a['_id'] = str(a['_id'])
            # Format: 12 Oct, 04:30 PM
            if a.get('created_at'):
                a['date'] = a['created_at'].strftime('%d %b, %I:%M %p')
            else:
                a['date'] = 'Just now'
        return jsonify(alerts)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/emergency', methods=['POST'])
@login_required
def add_emergency_alert():
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        data = clean_input_data(request.json)
        alert = {
            'patient_name': data.get('patient_name', 'Unknown'),
            'note': data.get('note', ''),
            'severity': data.get('severity', 'critical'), 
            'added_by': session.get('username', 'Staff'),
            'status': 'active',
            'created_at': datetime.now()
        }
        mongo.db.emergency_alerts.insert_one(alert)
        return jsonify({"message": "Alert added"}), 201
    except Exception as e:
        print(f"Emergency Save Error: {e}") # Added debug print
        return jsonify({"error": str(e)}), 500

@app.route('/api/emergency/<id>', methods=['DELETE'])
@login_required
def delete_emergency_alert(id):
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        # Instead of deleting, mark as resolved
        mongo.db.emergency_alerts.update_one(
            {'_id': ObjectId(id)},
            {'$set': {'status': 'resolved', 'resolved_at': datetime.now(), 'resolved_by': session.get('username', 'Staff')}}
        )
        return jsonify({"message": "Alert marked as resolved"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/patients/<id>/payment_history', methods=['GET'])
@role_required(['Admin'])
def get_patient_payment_history(id):
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        # 1. Get Patient Details to find the name
        patient = mongo.db.patients.find_one({'_id': ObjectId(id)})
        if not patient:
            return jsonify([])

        target_name = patient.get('name', '').strip().lower()
        target_id_str = str(id)
        
        # 2. Fetch ALL "Patient Fee" expenses (Incoming only)
        # We fetch all candidates first, then filter in Python for 100% accuracy matching your other API
        cursor = mongo.db.expenses.find({
            'type': 'incoming',
            'category': 'Patient Fee'
        }).sort('date', 1)
        
        history = []
        
        for doc in cursor:
            # --- MATCHING LOGIC ---
            is_match = False
            
            # Check A: Explicit ID Match (if available)
            doc_p_id = str(doc.get('patient_id', ''))
            if doc_p_id == target_id_str:
                is_match = True
            
            # Check B: Name Match in Note (The logic from your working API)
            # note format: "Partial payment from [Name] via..."
            if not is_match:
                note = doc.get('note', '').lower()
                if target_name and f"from {target_name}" in note:
                    is_match = True
            
            if is_match:
                # Safe date formatting
                date_str = '-'
                if doc.get('date'):
                    if isinstance(doc['date'], str):
                        date_str = doc['date'][:10]
                    else:
                        date_str = doc['date'].strftime('%d-%b-%Y')

                history.append({
                    'date': date_str,
                    'amount': doc.get('amount', 0),
                    'method': doc.get('payment_method', 'Cash'),
                    'note': doc.get('note', '')
                })
        return jsonify(history)

    except Exception as e:
        print(f"History error: {e}")
        return jsonify([])    

# --- OLD BALANCE / RECOVERY ROUTES ---

@app.route('/api/old-balances', methods=['GET'])
@role_required(['Admin'])
def get_old_balances():
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        month = request.args.get('month', type=int)
        year = request.args.get('year', type=int)

        query = {}
        if month and year:
            query = {
                '$or': [
                    {'month': month, 'year': year},
                    {
                        'month': {'$exists': False},
                        'year': {'$exists': False},
                        'created_at': {
                            '$gte': datetime(year, month, 1),
                            '$lt': datetime(year + 1, 1, 1) if month == 12 else datetime(year, month + 1, 1)
                        }
                    }
                ]
            }

        cursor = mongo.db.old_balances.find(query).sort('created_at', -1)
        balances = []
        for b in cursor:
            balances.append({
                'id': str(b['_id']),
                'name': b.get('name', ''),
                'amount': b.get('amount', 0),
                'commitment_date': b.get('commitment_date', ''),
                'last_call_date': b.get('last_call_date', ''),
                'note': b.get('note', '')
            })
        return jsonify(balances)
    except Exception as e:
        print(f"Old Balance Fetch Error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/old-balances', methods=['POST'])
@role_required(['Admin'])
def add_old_balance():
    if not check_db(): return jsonify({"error": "Database error"}), 500
    data = clean_input_data(request.json)
    try:
        month = int(data.get('month') or datetime.now().month)
        year = int(data.get('year') or datetime.now().year)
        record = {
            'name': data.get('name'),
            'amount': int(data.get('amount', 0)),
            'commitment_date': data.get('commitment_date'),
            'last_call_date': data.get('last_call_date'),
            'note': data.get('note', ''),
            'month': month,
            'year': year,
            'created_at': datetime.now(),
            'added_by': session.get('username')
        }
        result = mongo.db.old_balances.insert_one(record)
        return jsonify({"message": "Record added", "id": str(result.inserted_id)}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/old-balances/<id>', methods=['DELETE'])
@role_required(['Admin'])
def delete_old_balance(id):
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        mongo.db.old_balances.delete_one({'_id': ObjectId(id)})
        return jsonify({"message": "Record deleted"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# --- MANUAL DISCHARGE RECEIPT ROUTES ---

def _safe_int_amount(raw_val):
    try:
        return int(float(str(raw_val or '0').replace(',', '').strip() or '0'))
    except Exception:
        return 0

def _serialize_manual_receipt(doc):
    return {
        'id': str(doc.get('_id')),
        'patient_id': str(doc.get('patient_id')) if doc.get('patient_id') else '',
        'patient_name': doc.get('patient_name', ''),
        'father_name': doc.get('father_name', ''),
        'age': doc.get('age', ''),
        'cnic': doc.get('cnic', ''),
        'contact_no': doc.get('contact_no', ''),
        'area': doc.get('area', ''),
        'address': doc.get('address', ''),
        'admission_date': doc.get('admission_date', ''),
        'discharge_date': doc.get('discharge_date', ''),
        'stay_days': doc.get('stay_days', 0),
        'monthly_fee': doc.get('monthly_fee', 0),
        'fee_amount': doc.get('fee_amount', 0),
        'rehab_next_month_amount': doc.get('rehab_next_month_amount', 0),
        'test_amount': doc.get('test_amount', 0),
        'canteen_amount': doc.get('canteen_amount', 0),
        'laundry_amount': doc.get('laundry_amount', 0),
        'barbar_amount': doc.get('barbar_amount', 0),
        'medicine_amount': doc.get('medicine_amount', 0),
        'other_amount': doc.get('other_amount', 0),
        'received_amount': doc.get('received_amount', 0),
        'net_balance': doc.get('net_balance', 0),
        'notes': doc.get('notes', ''),
        'created_by': doc.get('created_by', ''),
        'updated_by': doc.get('updated_by', ''),
        'created_at': doc.get('created_at').isoformat() if doc.get('created_at') else '',
        'updated_at': doc.get('updated_at').isoformat() if doc.get('updated_at') else ''
    }

@app.route('/api/manual-discharge-receipts', methods=['GET'])
@role_required(['Admin'])
def list_manual_discharge_receipts():
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        q = (request.args.get('q') or '').strip().lower()
        cursor = mongo.db.manual_discharge_receipts.find().sort('created_at', -1)
        rows = []
        for doc in cursor:
            row = _serialize_manual_receipt(doc)
            if q:
                hay = f"{row.get('patient_name','')} {row.get('father_name','')} {row.get('contact_no','')} {row.get('cnic','')}".lower()
                if q not in hay:
                    continue
            rows.append(row)
        return jsonify(rows)
    except Exception as e:
        print(f"Manual receipt list error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/manual-discharge-receipts/<id>', methods=['GET'])
@role_required(['Admin'])
def get_manual_discharge_receipt(id):
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        doc = mongo.db.manual_discharge_receipts.find_one({'_id': ObjectId(id)})
        if not doc:
            return jsonify({"error": "Record not found"}), 404
        return jsonify(_serialize_manual_receipt(doc))
    except Exception as e:
        print(f"Manual receipt get error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/manual-discharge-receipts', methods=['POST'])
@role_required(['Admin'])
def create_manual_discharge_receipt():
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        data = clean_input_data(request.json or {})
        now = datetime.now()

        fee_amount = _safe_int_amount(data.get('fee_amount'))
        rehab_next_month_amount = _safe_int_amount(data.get('rehab_next_month_amount'))
        test_amount = _safe_int_amount(data.get('test_amount'))
        canteen_amount = _safe_int_amount(data.get('canteen_amount'))
        laundry_amount = _safe_int_amount(data.get('laundry_amount'))
        barbar_amount = _safe_int_amount(data.get('barbar_amount'))
        medicine_amount = _safe_int_amount(data.get('medicine_amount'))
        other_amount = _safe_int_amount(data.get('other_amount'))
        received_amount = _safe_int_amount(data.get('received_amount'))
        gross_total = fee_amount + rehab_next_month_amount + test_amount + canteen_amount + laundry_amount + barbar_amount + medicine_amount + other_amount
        net_balance = gross_total - received_amount

        patient_id = data.get('patient_id') or ''
        if patient_id and ObjectId.is_valid(patient_id):
            patient_id = ObjectId(patient_id)
        else:
            patient_id = None

        payload = {
            'patient_id': patient_id,
            'patient_name': data.get('patient_name', ''),
            'father_name': data.get('father_name', ''),
            'age': data.get('age', ''),
            'cnic': data.get('cnic', ''),
            'contact_no': data.get('contact_no', ''),
            'area': data.get('area', ''),
            'address': data.get('address', ''),
            'admission_date': data.get('admission_date', ''),
            'discharge_date': data.get('discharge_date', ''),
            'stay_days': _safe_int_amount(data.get('stay_days')),
            'monthly_fee': _safe_int_amount(data.get('monthly_fee')),
            'fee_amount': fee_amount,
            'rehab_next_month_amount': rehab_next_month_amount,
            'test_amount': test_amount,
            'canteen_amount': canteen_amount,
            'laundry_amount': laundry_amount,
            'barbar_amount': barbar_amount,
            'medicine_amount': medicine_amount,
            'other_amount': other_amount,
            'received_amount': received_amount,
            'net_balance': net_balance,
            'notes': data.get('notes', ''),
            'created_by': session.get('username', 'Admin'),
            'updated_by': session.get('username', 'Admin'),
            'created_at': now,
            'updated_at': now
        }

        result = mongo.db.manual_discharge_receipts.insert_one(payload)
        return jsonify({"message": "Manual discharge receipt saved", "id": str(result.inserted_id)}), 201
    except Exception as e:
        print(f"Manual receipt create error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/manual-discharge-receipts/<id>', methods=['PUT'])
@role_required(['Admin'])
def update_manual_discharge_receipt(id):
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        data = clean_input_data(request.json or {})
        fee_amount = _safe_int_amount(data.get('fee_amount'))
        rehab_next_month_amount = _safe_int_amount(data.get('rehab_next_month_amount'))
        test_amount = _safe_int_amount(data.get('test_amount'))
        canteen_amount = _safe_int_amount(data.get('canteen_amount'))
        laundry_amount = _safe_int_amount(data.get('laundry_amount'))
        barbar_amount = _safe_int_amount(data.get('barbar_amount'))
        medicine_amount = _safe_int_amount(data.get('medicine_amount'))
        other_amount = _safe_int_amount(data.get('other_amount'))
        received_amount = _safe_int_amount(data.get('received_amount'))
        gross_total = fee_amount + rehab_next_month_amount + test_amount + canteen_amount + laundry_amount + barbar_amount + medicine_amount + other_amount
        net_balance = gross_total - received_amount

        patient_id = data.get('patient_id') or ''
        if patient_id and ObjectId.is_valid(patient_id):
            patient_id = ObjectId(patient_id)
        else:
            patient_id = None

        payload = {
            'patient_id': patient_id,
            'patient_name': data.get('patient_name', ''),
            'father_name': data.get('father_name', ''),
            'age': data.get('age', ''),
            'cnic': data.get('cnic', ''),
            'contact_no': data.get('contact_no', ''),
            'area': data.get('area', ''),
            'address': data.get('address', ''),
            'admission_date': data.get('admission_date', ''),
            'discharge_date': data.get('discharge_date', ''),
            'stay_days': _safe_int_amount(data.get('stay_days')),
            'monthly_fee': _safe_int_amount(data.get('monthly_fee')),
            'fee_amount': fee_amount,
            'rehab_next_month_amount': rehab_next_month_amount,
            'test_amount': test_amount,
            'canteen_amount': canteen_amount,
            'laundry_amount': laundry_amount,
            'barbar_amount': barbar_amount,
            'medicine_amount': medicine_amount,
            'other_amount': other_amount,
            'received_amount': received_amount,
            'net_balance': net_balance,
            'notes': data.get('notes', ''),
            'updated_by': session.get('username', 'Admin'),
            'updated_at': datetime.now()
        }

        result = mongo.db.manual_discharge_receipts.update_one(
            {'_id': ObjectId(id)},
            {'$set': payload}
        )
        if result.matched_count == 0:
            return jsonify({"error": "Record not found"}), 404
        return jsonify({"message": "Manual discharge receipt updated"})
    except Exception as e:
        print(f"Manual receipt update error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/manual-discharge-receipts/<id>', methods=['DELETE'])
@role_required(['Admin'])
def delete_manual_discharge_receipt(id):
    if not check_db(): return jsonify({"error": "Database error"}), 500
    try:
        result = mongo.db.manual_discharge_receipts.delete_one({'_id': ObjectId(id)})
        if result.deleted_count == 0:
            return jsonify({"error": "Record not found"}), 404
        return jsonify({"message": "Record deleted"})
    except Exception as e:
        print(f"Manual receipt delete error: {e}")
        return jsonify({"error": str(e)}), 500

# --- HEALTH CHECK ENDPOINT (for cron-job.org) ---

@app.route('/health', methods=['GET'])
def health_check():
    """Lightweight health check endpoint for uptime monitoring"""
    return jsonify({"status": "ok", "timestamp": datetime.now().isoformat()}), 200

@app.route('/api/db-status', methods=['GET'])
def db_status():
    """Detailed database status for troubleshooting"""
    status = {
        "connected": check_db(),
        "has_mongo_obj": mongo is not None,
        "uri_configured": "MONGO_URI" in os.environ,
        "timestamp": datetime.now().isoformat()
    }
    if mongo:
        try:
            mongo.cx.admin.command('ping')
            status["ping"] = "pong"
        except Exception as e:
            status["ping_error"] = str(e)
    return jsonify(status)

@app.route('/ping', methods=['GET', 'HEAD'])
def ping():
    """Ultra-minimal ping endpoint - even lighter than /health"""
    return '', 200
    

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=True, host='0.0.0.0', port=port)
